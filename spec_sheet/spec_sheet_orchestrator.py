#!/usr/bin/env python3
"""
Spec Sheet Orchestrator
Main coordination class that orchestrates all spec sheet components
"""

import openpyxl
from typing import Dict, List, Any, Optional
from utils.config import JiraConfig
from utils.jira_client import JiraClient
from spec_sheet.settings.config_manager import ConfigManager
from spec_sheet.pricing.pricing_engine import PricingEngine
from spec_sheet.settings.risk_assessor import RiskAssessor
from spec_sheet.settings.moscow_manager import MoscowManager
from spec_sheet.spreadsheet.excel_manager import ExcelManager
from spec_sheet.sprint.sprint_planner import SprintPlanner


class SpecSheetOrchestrator:
    """Main orchestrator for coordinating all spec sheet operations"""
    
    def __init__(self):
        # Initialize configurations
        self.config_manager = ConfigManager()
        self.jira_config = JiraConfig()
        self.jira_client = JiraClient(self.jira_config)
        
        # Initialize specialized components
        self.pricing_engine = PricingEngine(self.config_manager)
        self.risk_assessor = RiskAssessor(self.config_manager, self.jira_config.type_of_work_field)
        self.moscow_manager = MoscowManager()
        self.excel_manager = ExcelManager(self.config_manager)
        self.sprint_planner = SprintPlanner(self.config_manager)
        
        # Initialize state
        self.selected_epics = None
        self.selected_version = None
        self.everything_else_items = None
        self.selected_moscow_priorities = None
    
    def test_connections(self) -> bool:
        """Test connections to Jira and spec sheet"""
        print("🔄 Testing connections...")
        
        # Test Jira connection
        if not self.jira_client.test_connection():
            return False
        
        # Test spec sheet access
        if not self.excel_manager.load_or_create_workbook():
            return False
        
        print("✅ All connections successful")
        return True
    
    def select_version_and_epics(self) -> bool:
        """Interactive selection of version/release and epics"""
        print("\n🔍 Select Version/Release to get Epics:")
        try:
            epics, selected_version = self.jira_client.get_epics_by_version_interactive()
            
            if not epics:
                print(f"⚠️  No epics found for '{selected_version}'. Please check your selection.")
                return False
            
            print(f"\n📊 Found {len(epics)} epic(s) for '{selected_version}'")
            
            # Get all issues for this version to find "Everything Else" items
            print(f"\n🔍 Finding additional items not tied to epics...")
            all_version_issues = self.jira_client.get_issues_for_version(
                selected_version if selected_version != "All Epics" else None,
                ['Story', 'Bug', 'Task', 'Subtask']  # Include various issue types
            )
            
            # Collect all story keys that are already linked to epics
            linked_story_keys = set()
            for epic in epics:
                epic_stories = self.jira_client.get_stories_for_epic(epic['key'])
                for story in epic_stories:
                    linked_story_keys.add(story['key'])
            
            # Find items not linked to any epic
            everything_else_items = []
            for issue in all_version_issues:
                if issue['key'] not in linked_story_keys:
                    everything_else_items.append(issue)
            
            print(f"📋 Found {len(everything_else_items)} additional items not tied to epics")
            
            # Store selected data
            self.selected_epics = epics
            self.selected_version = selected_version
            self.everything_else_items = everything_else_items
            
            return True
            
        except KeyboardInterrupt:
            print("\n⏹️  Operation cancelled.")
            return False
    
    def select_moscow_priorities(self) -> bool:
        """Interactive selection of MoSCoW priorities"""
        print("\n🎯 Filter by MoSCoW Priorities:")
        try:
            selected_priorities = self.moscow_manager.get_moscow_priorities_interactive()
            
            if selected_priorities is None:
                print("⏹️  Operation cancelled.")
                return False
            
            self.selected_moscow_priorities = selected_priorities
            return True
            
        except KeyboardInterrupt:
            print("\n⏹️  Selection cancelled. Using all priorities.")
            self.selected_moscow_priorities = ['Must Have', 'Should Have', 'Could Have', "Won't Have"]
            return True
    
    def sync_to_scope_sheet(self, sheet_name: str = 'Scope (Quantity)') -> bool:
        """Sync Jira data to the scope sheet with epic-story hierarchy"""
        print(f"🔄 Regenerating sheet: {sheet_name} from Jira data")
        
        # Use pre-selected epics and everything else items
        epics = self.selected_epics or []
        everything_else_items = self.everything_else_items or []
        version_info = self.selected_version or "All Epics"
        selected_priorities = self.selected_moscow_priorities
        
        if not epics and not everything_else_items:
            print("⚠️  No epics or other items found")
            return False
        
        # Clear and setup sheet
        if not self.excel_manager.clear_scope_sheet(sheet_name):
            return False
        
        ws = self.excel_manager.workbook[sheet_name]
        
        # Start adding data after headers (row 4)
        current_row = 4
        total_items = 0
        total_filtered_out = 0
        
        # Calculate totals for summary
        total_epics = len(epics)
        total_stories = 0
        total_story_points = 0
        costs = {'proven': 0, 'experimental_min': 0, 'experimental_max': 0, 'dependant': 0}
        
        # Process epics and their stories
        print(f"\n📊 Processing {len(epics)} epics...")
        for epic in epics:
            epic_key = epic['key']
            epic_summary = epic['fields']['summary']
            
            print(f"\n🔵 Epic: {epic_key} - {epic_summary}")
            
            # Get stories for this epic
            stories = self.jira_client.get_stories_for_epic(epic_key)
            
            if not stories:
                print(f"   ⚠️  No stories found for epic {epic_key}")
                continue
            
            # Filter stories by MoSCoW priorities if specified
            if selected_priorities and len(selected_priorities) < 4:
                filtered_stories, priority_counts = self.moscow_manager.filter_stories_by_moscow(stories, selected_priorities)
                filtered_out_count = len(stories) - len(filtered_stories)
                total_filtered_out += filtered_out_count
            else:
                filtered_stories = stories
            
            if not filtered_stories:
                print(f"   ⚠️  No stories match selected priorities for epic {epic_key}")
                continue
            
            # Add epic header
            self.excel_manager.add_epic_header(ws, current_row, epic_key, epic_summary)
            current_row += 1
            
            # Add stories under this epic
            for story in filtered_stories:
                story_key = story['key']
                story_summary = story['fields']['summary']
                story_points = self.jira_client.get_story_points(story) or 0
                
                # Determine risk profile and pricing
                risk_profile = self.risk_assessor.determine_risk_profile(story)
                moscow_priority = self.moscow_manager.get_moscow_priority(story)
                prices = self.pricing_engine.calculate_prices(story_points, risk_profile)
                
                print(f"   - {story_key}: {story_points} pts, {risk_profile}, {moscow_priority}")
                
                # Add story row
                self.excel_manager.add_story_row(ws, current_row, {
                    'key': story_key,
                    'summary': story_summary,
                    'story_points': story_points,
                    'moscow': moscow_priority,
                    'risk_profile': risk_profile,
                    'prices': prices
                })
                
                # Update totals
                total_stories += 1
                total_story_points += story_points
                
                if risk_profile == 'proven':
                    costs['proven'] += prices.get('fixed', 0)
                elif risk_profile == 'experimental':
                    costs['experimental_min'] += prices.get('minimum', 0)
                    costs['experimental_max'] += prices.get('maximum', 0)
                elif risk_profile == 'dependant':
                    costs['dependant'] += prices.get('hourly_estimate', 0)
                
                current_row += 1
                total_items += 1
            
            # Add spacing after each epic
            current_row += 1
        
        # Process "Everything Else" items not tied to epics
        if everything_else_items:
            print(f"\n📊 Processing {len(everything_else_items)} loose items...")
            
            # Filter everything else items by MoSCoW priorities if specified
            if selected_priorities and len(selected_priorities) < 4:
                filtered_everything_else, priority_counts = self.moscow_manager.filter_stories_by_moscow(everything_else_items, selected_priorities)
                filtered_out_count = len(everything_else_items) - len(filtered_everything_else)
                total_filtered_out += filtered_out_count
            else:
                filtered_everything_else = everything_else_items
            
            if filtered_everything_else:
                # Add "Everything Else" section header
                self.excel_manager.add_section_header(ws, current_row, "Everything Else (Not tied to epics)")
                current_row += 1
                
                for item in filtered_everything_else:
                    item_key = item['key']
                    item_summary = item['fields']['summary']
                    story_points = self.jira_client.get_story_points(item) or 0
                    item_type = item['fields'].get('issuetype', {}).get('name', 'Issue')
                    
                    # Determine risk profile and pricing
                    risk_profile = self.risk_assessor.determine_risk_profile(item)
                    moscow_priority = self.moscow_manager.get_moscow_priority(item)
                    prices = self.pricing_engine.calculate_prices(story_points, risk_profile)
                    
                    print(f"   - {item_key} ({item_type}): {story_points} pts, {risk_profile}, {moscow_priority}")
                    
                    # Add item row with type prefix
                    self.excel_manager.add_story_row(ws, current_row, {
                        'key': item_key,
                        'summary': f"[{item_type}] {item_summary}",
                        'story_points': story_points,
                        'moscow': moscow_priority,
                        'risk_profile': risk_profile,
                        'prices': prices
                    })
                    
                    # Update totals
                    total_stories += 1
                    total_story_points += story_points
                    
                    if risk_profile == 'proven':
                        costs['proven'] += prices.get('fixed', 0)
                    elif risk_profile == 'experimental':
                        costs['experimental_min'] += prices.get('minimum', 0)
                        costs['experimental_max'] += prices.get('maximum', 0)
                    elif risk_profile == 'dependant':
                        costs['dependant'] += prices.get('hourly_estimate', 0)
                    
                    current_row += 1
                    total_items += 1
                
                # Add spacing after everything else section
                current_row += 1
        
        # Add summary section at the end
        if total_items > 0:
            costs['grand_total'] = costs['proven'] + costs['experimental_max'] + costs['dependant']
            
            summary_data = {
                'version_info': version_info,
                'moscow_filter': selected_priorities,
                'total_epics': total_epics,
                'total_stories': total_stories,
                'total_story_points': total_story_points,
                'costs': costs
            }
            
            self.excel_manager.add_summary_section(ws, current_row + 1, summary_data)
        
        # Save the workbook
        self.excel_manager.save_workbook()
        
        # Report results
        filter_summary = f" (filtered out {total_filtered_out} items)" if total_filtered_out > 0 else ""
        print(f"\n✅ Sync completed! Generated {total_items} items in {sheet_name}{filter_summary}")
        
        return True
    
    def sync_all_sheets(self) -> bool:
        """Sync to all sheets - main entry point"""
        # Sync the main scope sheet
        success = self.sync_to_scope_sheet()
        
        # Add sprint planning analysis if main sync successful
        if success:
            try:
                print(f"\n🔄 Adding sprint planning analysis...")
                self.add_sprint_planning_sheet()
                print("✅ Sprint planning analysis added")
            except Exception as e:
                print(f"⚠️  Warning: Could not add sprint planning sheet: {e}")
        
        return success
    
    def add_sprint_planning_sheet(self) -> bool:
        """Add a sprint planning analysis sheet"""
        if not self.selected_epics:
            print("⚠️  No epics selected for sprint planning analysis")
            return False
        
        try:
            # Generate sprint planning data
            sprint_data = self.sprint_planner.create_sprint_planning_sheet_data(
                self.selected_epics,
                self.jira_client.get_stories_for_epic,
                self.jira_client.get_story_points,
                self.risk_assessor.determine_risk_profile
            )
            
            # Create or update sprint planning sheet
            sheet_name = 'Sprint Planning'
            
            # Remove existing sheet if it exists
            if sheet_name in self.excel_manager.workbook.sheetnames:
                self.excel_manager.workbook.remove(self.excel_manager.workbook[sheet_name])
            
            # Create new sheet
            ws = self.excel_manager.workbook.create_sheet(sheet_name)
            
            # Set up sprint planning sheet with data
            self._setup_sprint_planning_sheet(ws, sprint_data)
            
            # Save workbook
            self.excel_manager.save_workbook()
            print(f"✅ Added sprint planning analysis to {sheet_name} sheet")
            return True
            
        except Exception as e:
            print(f"❌ Error adding sprint planning sheet: {e}")
            return False
    
    def _setup_sprint_planning_sheet(self, ws, report: Dict[str, Any]):
        """Set up the sprint planning sheet with analysis data"""
        current_row = 1
        
        # Title
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = "Sprint Planning & Team Composition Analysis"
        title_cell.font = openpyxl.styles.Font(bold=True, size=16)
        current_row += 3
        
        # Project Summary
        ws.cell(row=current_row, column=1).value = "Project Overview"
        ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Total Story Points:"
        ws.cell(row=current_row, column=2).value = report['total_story_points']
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Number of Epics:"
        ws.cell(row=current_row, column=2).value = len(report['epic_breakdown'])
        current_row += 2
        
        # Epic Breakdown
        ws.cell(row=current_row, column=1).value = "Epic Breakdown"
        ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True, size=12)
        current_row += 1
        
        # Epic headers
        headers = ["Epic Key", "Summary", "Story Points", "Story Count"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=i)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
        current_row += 1
        
        # Epic data
        for epic_key, epic_data in report['epic_breakdown'].items():
            ws.cell(row=current_row, column=1).value = epic_key
            ws.cell(row=current_row, column=2).value = epic_data['summary']
            ws.cell(row=current_row, column=3).value = epic_data['story_points']
            ws.cell(row=current_row, column=4).value = epic_data['story_count']
            current_row += 1
        
        current_row += 2
        
        # Team Comparison
        ws.cell(row=current_row, column=1).value = "Team Composition Comparison"
        ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        current_row += 1
        
        # Team comparison headers
        team_headers = ["Team Size", "Total FTE", "Velocity (SP/sprint)", "Sprints Needed", "Weeks", "Months"]
        for i, header in enumerate(team_headers, 2):
            cell = ws.cell(row=current_row, column=i)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
        current_row += 1
        
        # Team comparison data
        for team_name, estimates in report['team_comparison'].items():
            team = self.sprint_planner.default_teams[team_name]
            ws.cell(row=current_row, column=1).value = team_name.title()
            ws.cell(row=current_row, column=2).value = len(team.members)
            ws.cell(row=current_row, column=3).value = f"{team.get_total_fte():.1f}"
            ws.cell(row=current_row, column=4).value = f"{estimates['team_velocity']:.1f}"
            ws.cell(row=current_row, column=5).value = estimates['sprints']
            ws.cell(row=current_row, column=6).value = estimates['weeks']
            ws.cell(row=current_row, column=7).value = f"{estimates['months']:.1f}"
            current_row += 1
        
        current_row += 2
        
        # Risk Analysis
        ws.cell(row=current_row, column=1).value = "Risk Profile Analysis"
        ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        current_row += 1
        
        # Risk headers
        risk_headers = ["Risk Profile", "Stories", "Story Points", "% of Stories", "% of Story Points"]
        for i, header in enumerate(risk_headers, 1):
            cell = ws.cell(row=current_row, column=i)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
        current_row += 1
        
        # Risk data
        for risk_profile in ['proven', 'experimental', 'dependant']:
            ws.cell(row=current_row, column=1).value = risk_profile.title()
            ws.cell(row=current_row, column=2).value = report['risk_breakdown']['counts'][risk_profile]
            ws.cell(row=current_row, column=3).value = report['risk_breakdown']['story_points'][risk_profile]
            ws.cell(row=current_row, column=4).value = f"{report['risk_breakdown']['percentages'][risk_profile]:.1f}%"
            ws.cell(row=current_row, column=5).value = f"{report['risk_breakdown']['story_point_percentages'][risk_profile]:.1f}%"
            current_row += 1
        
        current_row += 2
        
        # Recommendations
        if report['recommendations']:
            ws.cell(row=current_row, column=1).value = "Recommendations"
            ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True, size=14)
            current_row += 1
            
            for i, recommendation in enumerate(report['recommendations'], 1):
                ws.cell(row=current_row, column=1).value = f"{i}. {recommendation}"
                current_row += 1
        
        # Adjust column widths
        column_widths = [25, 50, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width 