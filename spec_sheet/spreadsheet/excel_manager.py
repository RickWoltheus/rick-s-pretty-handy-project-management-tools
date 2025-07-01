#!/usr/bin/env python3
"""
Excel Manager for Spec Sheet
Handles Excel spreadsheet operations, formatting, and content generation
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List, Any
import pandas as pd
import os
from spec_sheet.settings.config_manager import ConfigManager


class ExcelManager:
    """Manages Excel spreadsheet operations and formatting"""
    
    def __init__(self, config_manager: ConfigManager, spec_sheet_path: str = "spec-sheet.xlsx"):
        self.config_manager = config_manager
        self.spec_sheet_path = spec_sheet_path
        self.workbook = None
        self.settings_config = config_manager.settings_config
        self.dod_config = config_manager.dod_config
    
    def load_or_create_workbook(self) -> bool:
        """Load existing workbook or create a new one"""
        try:
            if os.path.exists(self.spec_sheet_path):
                self.workbook = openpyxl.load_workbook(self.spec_sheet_path)
                print("✅ Loaded existing spec sheet structure")
            else:
                print(f"📝 Spec sheet not found at {self.spec_sheet_path}, creating new one...")
                self._create_new_workbook()
                print("✅ Created new spec sheet with default structure")
            
            return True
            
        except Exception as e:
            print(f"❌ Error with spec sheet: {e}")
            return False
    
    def _create_new_workbook(self):
        """Create a new workbook with all required sheets"""
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Create required sheets
        self._create_scope_sheet()
        self._create_dod_sheet()
        self._create_settings_sheet()
        
        # Save the new workbook
        self.save_workbook()
    
    def _create_scope_sheet(self):
        """Create Scope (Quantity) sheet"""
        scope_sheet = self.workbook.create_sheet('Scope (Quantity)')
        self._setup_scope_headers(scope_sheet)
    
    def _create_dod_sheet(self):
        """Create Definition of Done sheet with default values"""
        dod_sheet = self.workbook.create_sheet('Definition of Done (Quality)')
        self._setup_dod_sheet(dod_sheet)
    
    def _create_settings_sheet(self):
        """Create Settings sheet with default values"""
        settings_sheet = self.workbook.create_sheet('Settings')
        self._setup_settings_sheet(settings_sheet)
    
    def _setup_scope_headers(self, ws):
        """Set up the scope sheet header structure"""
        # Main header row 1 - complex header with scope description
        scope_header = ("Scope\n\nWhat is in scope (features, functionality, integrations) and what is out of scope. "
                       "We define risk profiles transparently. This visibility helps to allocate resources effectively, "
                       "avoid unmanageable scope creep, and ultimately create a shared commitment.\n\n"
                       "* Prices are all-in, meaning they include project management and end-to-end delivery of the product. "
                       "Prices are ex. VAT.")
        
        headers = [
            scope_header,
            "MoSCoW\n\nMust Have\nShould Have\nCould Have\nWon't Have (out of scope)",
            "Risk Profile\n\nLow (Proven)\nModerate (Experimental)\nHigh (Dependant)",
            "Notes\n\n",
            "Proven\nLow risk profile.\n\nPrice is fixed. Scope is clear and we own this work.",
            "Price\nFixed",
            "Experimental\nModerate risk profile.\n\nPrice can be 30% lower or higher but never more. We share the risk when the work is experimental.",
            "Price\nMinimum",
            "Price\nMaximum",
            "Dependant\nHigh risk profile.\n\nPrice is estimation, but will be pay-by-hour. We depend on to-be-delivered APIs or externally employed team members.",
            "Price\nEstimate, but will be pay-by-hour.\n\nOur rate on 05-02-2025 is €127,16/h - 25% discount."
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i).value = header
        
        # Style the headers
        self._apply_header_formatting(ws)
    
    def _apply_header_formatting(self, ws):
        """Apply formatting to headers"""
        header_font = Font(bold=True, size=9)
        for col in range(1, 12):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set column widths from configuration
        column_widths = self.settings_config["ui_formatting"]["default_column_widths"]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Set row height from configuration
        ws.row_dimensions[1].height = self.settings_config["ui_formatting"]["header_row_height"]
    
    def _setup_dod_sheet(self, ws):
        """Set up Definition of Done sheet using configuration"""
        headers = [
            "Definition of Done",
            "MoSCoW",
            "Price Impact",
            "Price Impact %"
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i).value = header
            ws.cell(row=1, column=i).font = Font(bold=True)
        
        current_row = 2
        
        # Use configuration to build DoD sheet
        for category in self.dod_config.get('categories', []):
            category_name = category.get('name', 'Unknown Category')
            
            # Add category header
            ws.cell(row=current_row, column=1).value = category_name
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            # Add items for this category
            for item in category.get('items', []):
                desc = item.get('description', '')
                moscow = item.get('moscow', 'Should Have')
                impact_desc = item.get('impact_points', '')
                impact_pct = item.get('impact_percentage', 0.0)
                
                ws.cell(row=current_row, column=1).value = desc
                ws.cell(row=current_row, column=2).value = moscow
                ws.cell(row=current_row, column=3).value = impact_desc
                ws.cell(row=current_row, column=4).value = impact_pct
                
                current_row += 1
            
            # Add empty row between categories
            current_row += 1
    
    def _setup_settings_sheet(self, ws):
        """Set up default Settings sheet using configurable values"""
        pricing_config = self.settings_config["pricing"]
        calculated_hourly_rate = self.config_manager.get_hourly_rate()
        
        settings_data = [
            ("Range for Experimental pricing (up & down)", pricing_config["experimental_variance"]),
            ("Base price of 8 story points", pricing_config["base_story_point_price"] * 8),
            ("Base price of 1 story point", pricing_config["base_story_point_price"]),
            ("Hourly rate", calculated_hourly_rate),
        ]
        
        ws.cell(row=1, column=1).value = "Settings"
        ws.cell(row=1, column=1).font = Font(bold=True)
        
        for i, (setting, value) in enumerate(settings_data, 2):
            ws.cell(row=i, column=1).value = setting
            ws.cell(row=i, column=2).value = value
    
    def clear_scope_sheet(self, sheet_name: str = 'Scope (Quantity)'):
        """Clear existing content and set up fresh headers"""
        if sheet_name not in self.workbook.sheetnames:
            print(f"❌ Sheet '{sheet_name}' not found in workbook")
            return False
        
        ws = self.workbook[sheet_name]
        
        # Clear all existing content
        ws.delete_rows(1, ws.max_row)
        
        # Set up fresh headers
        self._setup_scope_headers(ws)
        
        print("🧹 Cleared existing content and set up fresh headers")
        return True
    
    def add_epic_header(self, ws, row: int, epic_key: str, epic_summary: str):
        """Add an epic header row"""
        ws.cell(row=row, column=1).value = f"[EPIC] {epic_summary} ({epic_key})"
        
        # Style the epic header
        epic_font = Font(bold=True, color="000000")
        epic_color = self.settings_config["ui_formatting"]["epic_background_color"]
        epic_fill = PatternFill(start_color=epic_color, end_color=epic_color, fill_type="solid")
        
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.font = epic_font
            cell.fill = epic_fill
    
    def add_story_row(self, ws, row: int, story_data: Dict[str, Any]):
        """Add a story row with all pricing information"""
        # Column mapping based on the original structure
        ws.cell(row=row, column=1).value = f"  └─ {story_data['summary']} ({story_data['key']})"
        ws.cell(row=row, column=2).value = story_data['moscow']
        ws.cell(row=row, column=3).value = story_data['risk_profile'].title()
        ws.cell(row=row, column=4).value = f"Story Points: {story_data['story_points']}"
        
        # Add pricing based on risk profile
        prices = story_data['prices']
        
        if story_data['risk_profile'] == 'proven':
            ws.cell(row=row, column=5).value = story_data['story_points']  # Story points
            ws.cell(row=row, column=6).value = prices.get('fixed', 0)  # Fixed price
            
        elif story_data['risk_profile'] == 'experimental':
            ws.cell(row=row, column=7).value = story_data['story_points']  # Story points
            ws.cell(row=row, column=8).value = prices.get('minimum', 0)  # Min price
            ws.cell(row=row, column=9).value = prices.get('maximum', 0)  # Max price
            
        elif story_data['risk_profile'] == 'dependant':
            ws.cell(row=row, column=10).value = story_data['story_points']  # Story points
            ws.cell(row=row, column=11).value = prices.get('hourly_estimate', 0)  # Hourly estimate
    
    def add_section_header(self, ws, row: int, section_title: str):
        """Add a section header row for non-epic hierarchies"""
        ws.cell(row=row, column=1).value = f"[SECTION] {section_title}"
        
        # Style the section header
        section_font = Font(bold=True, color="000000")
        section_color = self.settings_config["ui_formatting"]["section_background_color"]
        section_fill = PatternFill(start_color=section_color, end_color=section_color, fill_type="solid")
        
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.font = section_font
            cell.fill = section_fill
    
    def add_summary_section(self, ws, start_row: int, summary_data: Dict[str, Any]):
        """Add a summary section with totals"""
        current_row = start_row + 1
        
        # Add summary header
        version_info = summary_data.get('version_info', 'All Epics')
        moscow_filter = summary_data.get('moscow_filter', [])
        
        if moscow_filter and len(moscow_filter) < 4:
            filter_text = f" ({', '.join(moscow_filter)} only)"
        else:
            filter_text = ""
            
        ws.cell(row=current_row, column=1).value = f"PROJECT SUMMARY - {version_info}{filter_text}"
        summary_font = Font(bold=True, size=12, color="FFFFFF")
        header_color = self.settings_config["ui_formatting"]["summary_header_color"]
        summary_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        
        for col in range(1, 12):
            cell = ws.cell(row=current_row, column=col)
            cell.font = summary_font
            cell.fill = summary_fill
        
        current_row += 2
        
        # Add summary data
        ws.cell(row=current_row, column=1).value = f"Total Epics: {summary_data.get('total_epics', 0)}"
        current_row += 1
        ws.cell(row=current_row, column=1).value = f"Total Stories: {summary_data.get('total_stories', 0)}"
        current_row += 1
        ws.cell(row=current_row, column=1).value = f"Total Story Points: {summary_data.get('total_story_points', 0)}"
        current_row += 2
        
        # Add cost summary
        costs = summary_data.get('costs', {})
        ws.cell(row=current_row, column=1).value = "COST SUMMARY:"
        current_row += 1
        
        if costs.get('proven', 0) > 0:
            ws.cell(row=current_row, column=1).value = "Proven (Fixed):"
            ws.cell(row=current_row, column=6).value = f"€{costs['proven']:,.2f}"
            current_row += 1
        
        if costs.get('experimental_min', 0) > 0:
            ws.cell(row=current_row, column=1).value = "Experimental (Range):"
            ws.cell(row=current_row, column=8).value = f"€{costs['experimental_min']:,.2f}"
            ws.cell(row=current_row, column=9).value = f"€{costs['experimental_max']:,.2f}"
            current_row += 1
        
        if costs.get('dependant', 0) > 0:
            ws.cell(row=current_row, column=1).value = "Dependant (Estimate):"
            ws.cell(row=current_row, column=11).value = f"€{costs['dependant']:,.2f}"
            current_row += 1
        
        # Add grand total estimate
        grand_total = costs.get('grand_total', 0)
        current_row += 1
        ws.cell(row=current_row, column=1).value = "ESTIMATED TOTAL (worst case):"
        ws.cell(row=current_row, column=11).value = f"€{grand_total:,.2f}"
        
        # Style the totals
        total_font = Font(bold=True, size=11)
        for row in range(start_row + 3, current_row + 1):
            ws.cell(row=row, column=1).font = total_font
    
    def save_workbook(self):
        """Save the workbook"""
        if self.workbook:
            self.workbook.save(self.spec_sheet_path)
            print(f"💾 Saved workbook to {self.spec_sheet_path}")
    
    def load_dod_impacts_from_sheet(self) -> Dict[str, float]:
        """Load Definition of Done impacts from the spec sheet"""
        dod_impacts = {}
        try:
            df = pd.read_excel(self.spec_sheet_path, sheet_name='Definition of Done (Quality)')
            
            for idx, row in df.iterrows():
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[3]):
                    description = str(row.iloc[0])
                    # Skip category headers and sum row
                    if description in ['Sum'] or description.endswith('&') or description.endswith('compatibility') or description.endswith('deployment'):
                        continue
                    
                    try:
                        impact = float(row.iloc[3])
                        # Create a clean key from the description
                        clean_key = description.lower().replace(' ', '_').replace('(', '').replace(')', '').replace(',', '').replace('-', '_')
                        dod_impacts[clean_key] = impact
                        print(f"   📋 Loaded DoD: {description} = {impact:.1%}")
                    except (ValueError, TypeError):
                        continue
                        
        except Exception as e:
            print(f"Warning: Could not load DoD impacts from sheet: {e}")
            return self.config_manager.get_dod_impacts()
            
        if not dod_impacts:
            print("Warning: No DoD impacts found in sheet, using defaults")
            return self.config_manager.get_dod_impacts()
            
        return dod_impacts 