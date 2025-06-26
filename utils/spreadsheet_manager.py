import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional
from .config import SpreadsheetConfig

class SpreadsheetManager:
    """Manager for Excel spreadsheet operations"""
    
    def __init__(self, config: SpreadsheetConfig):
        self.config = config
        self.workbook = None
        self.worksheet = None
    
    def load_or_create_workbook(self):
        """Load existing workbook or create a new one"""
        if os.path.exists(self.config.file_path):
            try:
                self.workbook = load_workbook(self.config.file_path)
                print(f"📊 Loaded existing spreadsheet: {self.config.file_path}")
            except Exception as e:
                print(f"Error loading existing spreadsheet: {e}")
                print("Creating new spreadsheet...")
                self._create_new_workbook()
        else:
            self._create_new_workbook()
        
        # Get or create the specified worksheet
        if self.config.sheet_name in self.workbook.sheetnames:
            self.worksheet = self.workbook[self.config.sheet_name]
        else:
            self.worksheet = self.workbook.create_sheet(self.config.sheet_name)
            self._setup_headers()
    
    def _create_new_workbook(self):
        """Create a new workbook with proper formatting"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = self.config.sheet_name
        self._setup_headers()
        print(f"📝 Created new spreadsheet: {self.config.file_path}")
    
    def _setup_headers(self):
        """Set up column headers with formatting"""
        headers = [
            "Epic/Story",
            "Task Description", 
            "Story Points",
            "Cost per Point",
            "Total Cost"
        ]
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths
        column_widths = [30, 50, 15, 15, 15]
        for col_num, width in enumerate(column_widths, 1):
            self.worksheet.column_dimensions[get_column_letter(col_num)].width = width
    
    def clear_data_rows(self):
        """Clear all data rows (keeping headers)"""
        if self.worksheet.max_row > 1:
            self.worksheet.delete_rows(self.config.data_start_row, self.worksheet.max_row)
    
    def add_epic_row(self, epic_key: str, epic_summary: str, row_num: int):
        """Add an epic row with formatting"""
        epic_cell = self.worksheet.cell(row=row_num, column=self.config.epic_column)
        epic_cell.value = f"[EPIC] {epic_summary} ({epic_key})"
        
        # Epic styling
        epic_font = Font(bold=True, color="000000")
        epic_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        epic_cell.font = epic_font
        epic_cell.fill = epic_fill
        
        # Apply styling to the entire row
        for col in range(1, 6):
            cell = self.worksheet.cell(row=row_num, column=col)
            cell.fill = epic_fill
    
    def add_story_row(self, story_key: str, story_summary: str, story_points: Optional[float], row_num: int):
        """Add a story row with cost calculations"""
        # Story description
        story_cell = self.worksheet.cell(row=row_num, column=self.config.story_column)
        story_cell.value = f"  └─ {story_summary} ({story_key})"
        
        # Story points
        points_cell = self.worksheet.cell(row=row_num, column=self.config.story_points_column)
        points_cell.value = story_points if story_points is not None else 0
        
        # Cost per point
        cost_per_point_cell = self.worksheet.cell(row=row_num, column=self.config.cost_per_point_column)
        cost_per_point_cell.value = self.config.cost_per_story_point
        
        # Total cost calculation
        total_cost_cell = self.worksheet.cell(row=row_num, column=self.config.total_cost_column)
        if story_points is not None:
            points_col = get_column_letter(self.config.story_points_column)
            cost_col = get_column_letter(self.config.cost_per_point_column)
            total_cost_cell.value = f"={points_col}{row_num}*{cost_col}{row_num}"
        else:
            total_cost_cell.value = 0
    
    def add_epic_summary_row(self, epic_key: str, total_points: float, row_num: int):
        """Add a summary row for epic totals"""
        summary_cell = self.worksheet.cell(row=row_num, column=self.config.story_column)
        summary_cell.value = f"  EPIC TOTAL:"
        
        points_cell = self.worksheet.cell(row=row_num, column=self.config.story_points_column)
        points_cell.value = total_points
        
        cost_per_point_cell = self.worksheet.cell(row=row_num, column=self.config.cost_per_point_column)
        cost_per_point_cell.value = self.config.cost_per_story_point
        
        total_cost_cell = self.worksheet.cell(row=row_num, column=self.config.total_cost_column)
        total_cost_cell.value = total_points * self.config.cost_per_story_point
        
        # Summary styling
        summary_font = Font(bold=True, italic=True)
        summary_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        for col in range(1, 6):
            cell = self.worksheet.cell(row=row_num, column=col)
            cell.font = summary_font
            cell.fill = summary_fill
    
    def add_grand_total_row(self, total_points: float, total_cost: float, row_num: int):
        """Add a grand total row at the bottom"""
        # Empty row first
        row_num += 1
        
        # Grand total
        total_label_cell = self.worksheet.cell(row=row_num, column=self.config.story_column)
        total_label_cell.value = "GRAND TOTAL:"
        
        total_points_cell = self.worksheet.cell(row=row_num, column=self.config.story_points_column)
        total_points_cell.value = total_points
        
        total_cost_cell = self.worksheet.cell(row=row_num, column=self.config.total_cost_column)
        total_cost_cell.value = total_cost
        
        # Grand total styling
        grand_total_font = Font(bold=True, size=12, color="FFFFFF")
        grand_total_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in range(1, 6):
            cell = self.worksheet.cell(row=row_num, column=col)
            cell.font = grand_total_font
            cell.fill = grand_total_fill
    
    def save(self):
        """Save the workbook"""
        try:
            self.workbook.save(self.config.file_path)
            print(f"💾 Spreadsheet saved: {self.config.file_path}")
        except Exception as e:
            print(f"❌ Error saving spreadsheet: {e}")
            raise
    
    def get_workbook_info(self):
        """Get information about the current workbook"""
        if not self.worksheet:
            return "No worksheet loaded"
        
        info = {
            'file_path': self.config.file_path,
            'sheet_name': self.worksheet.title,
            'rows': self.worksheet.max_row,
            'cols': self.worksheet.max_column
        }
        return info 