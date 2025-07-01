#!/usr/bin/env python3
"""
Tests for Timeline Generator functionality
"""

import pytest
import sys
import os
from datetime import date, timedelta
from unittest.mock import Mock, patch, MagicMock
import tempfile
import openpyxl

# Add project root to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from timeline.timeline_generator import (
    TimelineGenerator, TeamMemberInfo, HolidayInfo, Sprint,
    DutchHolidayAPI
)

class TestTeamMemberInfo:
    """Test TeamMemberInfo dataclass"""
    
    def test_team_member_creation(self):
        """Test basic team member creation"""
        member = TeamMemberInfo(
            name="Alice Johnson",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=8,
            hourly_rate=85
        )
        
        assert member.name == "Alice Johnson"
        assert member.role == "Developer"
        assert member.fte == 1.0
        assert member.story_points_per_sprint == 8
        assert member.hourly_rate == 85
        assert member.seniority_level == "Mid"  # Default
        assert member.start_date == date.today()  # Auto-set
        assert member.working_days == [0, 1, 2, 3, 4]  # Mon-Fri default
    
    def test_team_member_with_custom_values(self):
        """Test team member with custom values"""
        custom_start = date(2025, 1, 1)
        custom_end = date(2025, 12, 31)
        custom_holidays = [date(2025, 7, 15)]
        
        member = TeamMemberInfo(
            name="Bob Smith",
            role="Senior Developer",
            fte=0.8,
            story_points_per_sprint=10,
            hourly_rate=120,
            start_date=custom_start,
            end_date=custom_end,
            seniority_level="Senior",
            skills=["Python", "React"],
            custom_holidays=custom_holidays,
            working_days=[0, 1, 2, 3]  # Mon-Thu
        )
        
        assert member.fte == 0.8
        assert member.start_date == custom_start
        assert member.end_date == custom_end
        assert member.seniority_level == "Senior"
        assert member.skills == ["Python", "React"]
        assert member.custom_holidays == custom_holidays
        assert member.working_days == [0, 1, 2, 3]

class TestHolidayInfo:
    """Test HolidayInfo dataclass"""
    
    def test_national_holiday(self):
        """Test national holiday creation"""
        holiday = HolidayInfo(
            date=date(2025, 1, 1),
            name="New Year's Day",
            is_national=True
        )
        
        assert holiday.date == date(2025, 1, 1)
        assert holiday.name == "New Year's Day"
        assert holiday.is_national is True
        assert holiday.affected_members == []  # Affects everyone
    
    def test_custom_holiday(self):
        """Test custom holiday with specific members"""
        holiday = HolidayInfo(
            date=date(2025, 7, 15),
            name="Alice's Vacation",
            is_national=False,
            affected_members=["Alice Johnson"]
        )
        
        assert holiday.is_national is False
        assert holiday.affected_members == ["Alice Johnson"]

class TestDutchHolidayAPI:
    """Test Dutch holiday API functionality"""
    
    @patch('requests.get')
    def test_fetch_holidays_success(self, mock_get):
        """Test successful holiday fetching"""
        # Mock API response
        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.json.return_value = [
            {
                'date': '2025-01-01',
                'name': 'New Year\'s Day'
            },
            {
                'date': '2025-04-27',
                'name': 'King\'s Day'
            }
        ]
        mock_get.return_value = mock_response
        
        holidays = DutchHolidayAPI.fetch_holidays(2025)
        
        assert len(holidays) == 2
        assert holidays[0].date == date(2025, 1, 1)
        assert holidays[0].name == "New Year's Day"
        assert holidays[0].is_national is True
        assert holidays[1].date == date(2025, 4, 27)
        assert holidays[1].name == "King's Day"
    
    @patch('requests.get')
    def test_fetch_holidays_api_error(self, mock_get):
        """Test fallback when API fails"""
        mock_get.side_effect = Exception("API Error")
        
        holidays = DutchHolidayAPI.fetch_holidays(2025)
        
        # Should get fallback holidays
        assert len(holidays) > 0
        # Check for New Year's Day in fallback
        new_years = [h for h in holidays if h.date.month == 1 and h.date.day == 1]
        assert len(new_years) == 1

class TestTimelineGenerator:
    """Test TimelineGenerator main class"""
    
    def test_init(self):
        """Test timeline generator initialization"""
        timeline = TimelineGenerator()
        
        assert timeline.team_members == []
        assert timeline.holidays == []
        assert timeline.sprints == []
        assert timeline.project_start_date is None
        assert timeline.project_end_date is None
    
    def test_add_team_member(self):
        """Test adding team members"""
        timeline = TimelineGenerator()
        
        member = TeamMemberInfo(
            name="Alice Johnson",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=8,
            hourly_rate=85
        )
        
        timeline.add_team_member(member)
        
        assert len(timeline.team_members) == 1
        assert timeline.team_members[0] == member
    
    def test_add_custom_holiday(self):
        """Test adding custom holidays"""
        timeline = TimelineGenerator()
        
        holiday = HolidayInfo(
            date=date(2025, 7, 15),
            name="Company Holiday",
            is_national=False
        )
        
        timeline.add_custom_holiday(holiday)
        
        assert len(timeline.holidays) == 1
        assert timeline.holidays[0] == holiday

    def test_add_holiday_range(self):
        """Test adding holiday ranges"""
        timeline = TimelineGenerator()
        
        # Add a 3-day holiday range
        holidays_added = timeline.add_holiday_range(
            start_date=date(2025, 7, 15),
            end_date=date(2025, 7, 17),
            name="Summer Break",
            is_national=False,
            affected_members=["Alice"]
        )
        
        assert holidays_added == 3
        assert len(timeline.holidays) == 3
        
        # Check each day was added
        expected_dates = [date(2025, 7, 15), date(2025, 7, 16), date(2025, 7, 17)]
        actual_dates = [h.date for h in timeline.holidays]
        assert actual_dates == expected_dates
        
        # Check all holidays have correct properties
        for holiday in timeline.holidays:
            assert "Summer Break" in holiday.name
            assert holiday.is_national is False
            assert holiday.affected_members == ["Alice"]

    def test_add_workdays_holiday_range(self):
        """Test adding workdays-only holiday ranges"""
        timeline = TimelineGenerator()
        
        # Add range that includes a weekend (Mon Jul 14 to Sun Jul 20, 2025)
        holidays_added = timeline.add_workdays_holiday_range(
            start_date=date(2025, 7, 14),  # Monday
            end_date=date(2025, 7, 20),    # Sunday
            name="Work Week Break",
            is_national=False
        )
        
        # Should only add Mon-Fri (5 working days)
        assert holidays_added == 5
        assert len(timeline.holidays) == 5
        
        # Check only weekdays were added
        for holiday in timeline.holidays:
            assert holiday.date.weekday() < 5  # Monday=0, Friday=4
            assert "Work Week Break" in holiday.name

    def test_add_holiday_range_single_day(self):
        """Test adding a single day range"""
        timeline = TimelineGenerator()
        
        holidays_added = timeline.add_holiday_range(
            start_date=date(2025, 7, 15),
            end_date=date(2025, 7, 15),
            name="Single Day Holiday"
        )
        
        assert holidays_added == 1
        assert len(timeline.holidays) == 1
        assert timeline.holidays[0].name == "Single Day Holiday"

    def test_add_holiday_range_invalid_dates(self):
        """Test adding holiday range with invalid date order"""
        timeline = TimelineGenerator()
        
        # Start date after end date should fail gracefully
        result = timeline.add_holiday_range(
            start_date=date(2025, 7, 20),
            end_date=date(2025, 7, 15),
            name="Invalid Range"
        )
        
        assert result is None
        assert len(timeline.holidays) == 0
    
    def test_calculate_working_days(self):
        """Test working days calculation"""
        timeline = TimelineGenerator()
        
        # Add a holiday
        timeline.add_custom_holiday(HolidayInfo(
            date=date(2025, 7, 15),
            name="Test Holiday",
            is_national=True
        ))
        
        member = TeamMemberInfo(
            name="Alice",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=8,
            hourly_rate=85,
            working_days=[0, 1, 2, 3, 4]  # Mon-Fri
        )
        
        # Test a week with one holiday (July 14-18, 2025)
        # July 14 = Monday, July 15 = Tuesday (holiday), July 16-18 = Wed-Fri
        start_date = date(2025, 7, 14)
        end_date = date(2025, 7, 18)
        
        working_days = timeline.calculate_working_days(start_date, end_date, member)
        
        # Should be 4 days (Mon, Wed, Thu, Fri) - Tuesday is holiday
        assert working_days == 4
    
    def test_calculate_working_days_with_personal_holiday(self):
        """Test working days with personal holiday"""
        timeline = TimelineGenerator()
        
        # Personal holiday for specific member
        timeline.add_custom_holiday(HolidayInfo(
            date=date(2025, 7, 15),
            name="Alice's Vacation",
            is_national=False,
            affected_members=["Alice"]
        ))
        
        alice = TeamMemberInfo(
            name="Alice",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=8,
            hourly_rate=85
        )
        
        bob = TeamMemberInfo(
            name="Bob",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=8,
            hourly_rate=85
        )
        
        start_date = date(2025, 7, 14)  # Monday
        end_date = date(2025, 7, 18)    # Friday
        
        alice_working_days = timeline.calculate_working_days(start_date, end_date, alice)
        bob_working_days = timeline.calculate_working_days(start_date, end_date, bob)
        
        # Alice should have 4 working days (affected by holiday)
        assert alice_working_days == 4
        # Bob should have 5 working days (not affected)
        assert bob_working_days == 5
    
    def test_calculate_effective_capacity(self):
        """Test effective capacity calculation"""
        timeline = TimelineGenerator()
        
        member = TeamMemberInfo(
            name="Alice",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=10,
            hourly_rate=85
        )
        
        # Two week period with 10 working days (baseline)
        start_date = date(2025, 7, 14)
        end_date = date(2025, 7, 25)
        
        capacity = timeline.calculate_effective_capacity(member, start_date, end_date)
        
        # Should be close to 10 SP (10 working days out of 10 baseline)
        assert capacity == 10.0
    
    def test_calculate_effective_capacity_with_holiday(self):
        """Test capacity calculation with holiday impact"""
        timeline = TimelineGenerator()
        
        # Add a holiday during the sprint
        timeline.add_custom_holiday(HolidayInfo(
            date=date(2025, 7, 15),
            name="Test Holiday",
            is_national=True
        ))
        
        member = TeamMemberInfo(
            name="Alice",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=10,
            hourly_rate=85
        )
        
        # Two week period
        start_date = date(2025, 7, 14)
        end_date = date(2025, 7, 25)
        
        capacity = timeline.calculate_effective_capacity(member, start_date, end_date)
        
        # Should be less than 10 SP due to holiday
        assert capacity < 10.0
        assert capacity > 8.0  # But not too much less
    
    def test_generate_sprint_timeline_no_team(self):
        """Test sprint timeline generation with no team"""
        timeline = TimelineGenerator()
        
        sprints = timeline.generate_sprint_timeline(
            start_date=date(2025, 7, 1),
            total_story_points=100
        )
        
        assert sprints == []
    
    def test_generate_sprint_timeline_basic(self):
        """Test basic sprint timeline generation"""
        timeline = TimelineGenerator()
        
        # Add team member
        timeline.add_team_member(TeamMemberInfo(
            name="Alice",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=10,
            hourly_rate=85
        ))
        
        sprints = timeline.generate_sprint_timeline(
            start_date=date(2025, 7, 1),
            total_story_points=25,
            sprint_length_weeks=2
        )
        
        assert len(sprints) >= 2  # Should need at least 2 sprints for 25 SP
        assert sum(s.story_points for s in sprints) == 25
        
        # Check sprint structure
        first_sprint = sprints[0]
        assert first_sprint.number == 1
        assert first_sprint.name == "Sprint 1"
        assert first_sprint.start_date == date(2025, 7, 1)
        assert first_sprint.story_points <= first_sprint.team_velocity
    
    def test_load_story_points_from_spec_sheet(self):
        """Test loading story points from spec sheet - test the fallback case"""
        timeline = TimelineGenerator()
        
        # Test with non-existent file - should return 0
        story_points = timeline.load_story_points_from_spec_sheet("nonexistent.xlsx")
        assert story_points == 0
    
    def test_load_story_points_missing_file(self):
        """Test loading story points from missing file"""
        timeline = TimelineGenerator()
        
        story_points = timeline.load_story_points_from_spec_sheet("nonexistent.xlsx")
        
        assert story_points == 0
    
    def test_create_team_sheet(self):
        """Test Excel workbook creation"""
        timeline = TimelineGenerator()
        
        # Add sample data
        timeline.add_team_member(TeamMemberInfo(
            name="Alice",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=8,
            hourly_rate=85
        ))
        
        timeline.add_custom_holiday(HolidayInfo(
            date=date(2025, 7, 15),
            name="Test Holiday",
            is_national=False
        ))
        
        timeline.generate_sprint_timeline(
            start_date=date(2025, 7, 1),
            total_story_points=20
        )
        
        wb = timeline.create_team_sheet()
        
        # Check that all expected sheets exist
        expected_sheets = ["Team Members", "Holidays", "Sprint Timeline", "Calendar View"]
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames
        
        # Check team members sheet has data
        team_sheet = wb["Team Members"]
        assert team_sheet.cell(row=1, column=1).value == "Name"  # Header
        assert team_sheet.cell(row=2, column=1).value == "Alice"  # Data
    
    def test_save_timeline_workbook(self):
        """Test saving timeline workbook"""
        timeline = TimelineGenerator()
        
        timeline.add_team_member(TeamMemberInfo(
            name="Alice",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=8,
            hourly_rate=85
        ))
        
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_filename = os.path.join(temp_dir, "test_timeline.xlsx")
            
            saved_filename = timeline.save_timeline_workbook(temp_filename)
            
            assert saved_filename == temp_filename
            assert os.path.exists(temp_filename)
            
            # Verify it's a valid Excel file
            wb = openpyxl.load_workbook(temp_filename)
            assert "Team Members" in wb.sheetnames

class TestIntegrationScenarios:
    """Integration tests for realistic scenarios"""
    
    def test_realistic_project_timeline(self):
        """Test a realistic project timeline scenario"""
        timeline = TimelineGenerator()
        
        # Add realistic team
        timeline.add_team_member(TeamMemberInfo(
            name="Senior Dev",
            role="Senior Developer",
            fte=1.0,
            story_points_per_sprint=10,
            hourly_rate=120,
            seniority_level="Senior",
            skills=["Python", "React", "AWS"]
        ))
        
        timeline.add_team_member(TeamMemberInfo(
            name="Mid Dev",
            role="Mid Developer",
            fte=1.0,
            story_points_per_sprint=7,
            hourly_rate=95,
            seniority_level="Mid",
            skills=["JavaScript", "Node.js"]
        ))
        
        timeline.add_team_member(TeamMemberInfo(
            name="Part-time Designer",
            role="Designer",
            fte=0.8,
            story_points_per_sprint=5,
            hourly_rate=85,
            seniority_level="Senior",
            skills=["UI/UX", "Figma"],
            working_days=[0, 1, 2, 3]  # Mon-Thu
        ))
        
        # Add holidays
        timeline.add_custom_holiday(HolidayInfo(
            date=date(2025, 7, 21),
            name="Summer Break",
            is_national=False
        ))
        
        timeline.add_custom_holiday(HolidayInfo(
            date=date(2025, 8, 15),
            name="Senior Dev Vacation",
            is_national=False,
            affected_members=["Senior Dev"]
        ))
        
        # Generate timeline
        sprints = timeline.generate_sprint_timeline(
            start_date=date(2025, 7, 1),
            total_story_points=120,
            sprint_length_weeks=2
        )
        
        assert len(sprints) > 0
        assert abs(sum(s.story_points for s in sprints) - 120) < 0.1  # Allow for floating point precision
        
        # Verify timeline accounts for holidays and part-time team member
        total_capacity = sum(s.team_velocity for s in sprints)
        
        # Should have reduced capacity during holiday periods
        # and account for part-time designer
        expected_min_capacity = 15 * len(sprints)  # Conservative estimate
        expected_max_capacity = 25 * len(sprints)  # Optimistic estimate
        
        assert expected_min_capacity <= total_capacity <= expected_max_capacity
    
    def test_holiday_impact_analysis(self):
        """Test timeline with significant holiday impact"""
        timeline = TimelineGenerator()
        
        timeline.add_team_member(TeamMemberInfo(
            name="Alice",
            role="Developer",
            fte=1.0,
            story_points_per_sprint=10,
            hourly_rate=85
        ))
        
        # Add many holidays to see impact
        holiday_dates = [
            date(2025, 7, 1),   # Start of project
            date(2025, 7, 15),  # Mid first sprint
            date(2025, 7, 29),  # Second sprint
            date(2025, 8, 12),  # Third sprint
        ]
        
        for i, holiday_date in enumerate(holiday_dates):
            timeline.add_custom_holiday(HolidayInfo(
                date=holiday_date,
                name=f"Holiday {i+1}",
                is_national=True
            ))
        
        sprints = timeline.generate_sprint_timeline(
            start_date=date(2025, 7, 1),
            total_story_points=50
        )
        
        # Should generate more sprints due to holiday impact
        assert len(sprints) >= 5  # More sprints needed due to holidays
        
        # Each sprint should have reduced capacity
        for sprint in sprints:
            assert sprint.story_points <= sprint.team_velocity

if __name__ == "__main__":
    pytest.main([__file__]) 