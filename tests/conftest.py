"""
Pytest configuration and fixtures for Jira Spec Sheet Sync tests
"""
import pytest
import os
import tempfile
import shutil
from unittest.mock import Mock, patch
from typing import Dict, List
import openpyxl

from tests.test_data import (
    MOCK_EPICS, MOCK_STORIES, MOCK_VERSIONS, MOCK_FIELDS, TEST_CONFIG,
    get_mock_stories_for_epic, get_mock_epic_by_key
)

# Test fixtures

@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files"""
    temp_dir = tempfile.mkdtemp()
    yield temp_dir
    shutil.rmtree(temp_dir)

@pytest.fixture
def test_config():
    """Mock configuration for testing"""
    with patch.dict(os.environ, TEST_CONFIG):
        from utils.config import JiraConfig
        yield JiraConfig()

@pytest.fixture
def mock_jira_client(test_config):
    """Mock JIRA client with test data"""
    from utils.jira_client import JiraClient
    
    client = JiraClient(test_config)
    
    # Mock the API requests
    client.get_epics = Mock(return_value=MOCK_EPICS)
    client.get_stories_for_epic = Mock(side_effect=get_mock_stories_for_epic)
    client.get_story_points = Mock(side_effect=lambda story: story["fields"].get("customfield_10016", 0))
    client.get_project_versions = Mock(return_value=MOCK_VERSIONS["values"])
    client.get_available_versions = Mock(return_value=["v3", "v2"])
    client.get_version_details = Mock(side_effect=lambda name: next(
        (v for v in MOCK_VERSIONS["values"] if v["name"] == name), None
    ))
    client.test_connection = Mock(return_value=True)
    client._make_request = Mock(side_effect=_mock_api_request)
    
    return client

def _mock_api_request(endpoint: str, params: Dict = None) -> Dict:
    """Mock API request handler"""
    if 'search' in endpoint:
        # Mock JQL search
        jql = params.get('jql', '') if params else ''
        if 'issuetype = Epic' in jql:
            return {"issues": MOCK_EPICS}
        else:
            # Return stories based on epic key in JQL
            for epic_key in MOCK_STORIES.keys():
                if epic_key in jql:
                    return {"issues": MOCK_STORIES[epic_key]}
            return {"issues": []}
    
    elif 'version' in endpoint:
        return MOCK_VERSIONS
    
    elif 'field' in endpoint:
        return MOCK_FIELDS
    
    elif 'myself' in endpoint:
        return {"displayName": "Test User"}
    
    return {}

@pytest.fixture
def test_workbook(temp_dir):
    """Create a test Excel workbook"""
    wb = openpyxl.Workbook()
    
    # Create Scope sheet
    scope_ws = wb.active
    scope_ws.title = "Scope (Quantity)"
    
    # Add basic headers
    headers = [
        "Item", "MoSCoW", "Risk Profile", "Details", 
        "SP (Proven)", "Fixed Price", "SP (Experimental)", 
        "Min Price", "Max Price", "SP (Dependant)", "Hourly Estimate"
    ]
    for col, header in enumerate(headers, 1):
        scope_ws.cell(row=1, column=col, value=header)
    
    # Create Definition of Done (Quality) sheet with comprehensive structure
    dod_ws = wb.create_sheet("Definition of Done (Quality)")
    
    # Add headers
    headers = ["Definition of Done", "MoSCoW", "Price Impact", "Price Impact %"]
    for col, header in enumerate(headers, 1):
        dod_ws.cell(row=1, column=col, value=header)
    
    # Add comprehensive DoD data
    row_num = 2
    
    # Code quality & documentation
    dod_ws.cell(row=row_num, column=1, value="Code quality & documentation")
    row_num += 1
    code_quality_items = [
        ("Code is structured, modular, and follows best practices", "Must Have", "8", 0.04),
        ("Code is reviewed and approved via pull requests", "Must Have", "5", 0.03),
        ("Code is covered with relevant unit and integration tests", "Won't Have", "13", 0.0),
        ("Code is well-documented (inline comments, README, API docs)", "Could Have", "8", 0.04),
    ]
    for desc, moscow, impact_desc, impact_pct in code_quality_items:
        for col, value in enumerate([desc, moscow, impact_desc, impact_pct], 1):
            dod_ws.cell(row=row_num, column=col, value=value)
        row_num += 1
    row_num += 1  # Empty row
    
    # Performance & optimization
    dod_ws.cell(row=row_num, column=1, value="Performance & optimization")
    row_num += 1
    performance_items = [
        ("API calls are debounced to increase performance", "Must Have", "5", 0.03),
        ("Widget rendering is smooth with minimal performance impact", "Must Have", "13", 0.07),
        ("State management is efficient and avoids unnecessary re-renders", "Must Have", "8", 0.04),
    ]
    for desc, moscow, impact_desc, impact_pct in performance_items:
        for col, value in enumerate([desc, moscow, impact_desc, impact_pct], 1):
            dod_ws.cell(row=row_num, column=col, value=value)
        row_num += 1
    row_num += 1  # Empty row
    
    # Security & deployment (shortened for test)
    dod_ws.cell(row=row_num, column=1, value="Security & deployment")
    row_num += 1
    security_items = [
        ("Security vulnerabilities are identified and mitigated", "Must Have", "8", 0.04),
        ("Features are tested in multiple browsers and mobile devices", "Must Have", "13", 0.07),
    ]
    for desc, moscow, impact_desc, impact_pct in security_items:
        for col, value in enumerate([desc, moscow, impact_desc, impact_pct], 1):
            dod_ws.cell(row=row_num, column=col, value=value)
        row_num += 1
    row_num += 1  # Empty row
    
    # Sum row - calculate from actual values
    row_num += 1
    dod_ws.cell(row=row_num, column=3, value="Sum")
    
    # Calculate sum from the test DoD items
    test_sum = 0.04 + 0.03 + 0.0 + 0.04 + 0.03 + 0.07 + 0.04 + 0.04 + 0.07  # Sum of all impact percentages above
    dod_ws.cell(row=row_num, column=4, value=test_sum)
    
    # Create Settings sheet
    settings_ws = wb.create_sheet("Settings")
    settings_data = [
        ["Setting", "Value"],
        ["Base Story Point Price", 130],
        ["Experimental Variance", 0.3],
        ["Hourly Rate", 95.37],
        ["DoD Impact Total", 0.63]
    ]
    for row, data in enumerate(settings_data, 1):
        for col, value in enumerate(data, 1):
            settings_ws.cell(row=row, column=col, value=value)
    
    # Save workbook
    test_file = os.path.join(temp_dir, "test-spec-sheet.xlsx")
    wb.save(test_file)
    
    return test_file

@pytest.fixture
def mock_spec_sheet_sync(mock_jira_client, test_workbook, temp_dir):
    """Mock EnhancedSpecSheetSync instance for testing"""
    # Patch the class to use our mocks
    with patch('spec-sheet.spec-sheet-generator.JiraClient') as mock_client_class:
        with patch('spec-sheet.spec-sheet-generator.JiraConfig') as mock_config_class:
            mock_client_class.return_value = mock_jira_client
            mock_config_class.return_value = mock_jira_client.config
            
            # Import after patching
            import sys
            sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
            
            from spec_sheet.spec_sheet_generator import EnhancedSpecSheetSync
            
            # Create instance
            sync = EnhancedSpecSheetSync()
            sync.spec_sheet_path = test_workbook
            
            yield sync

@pytest.fixture
def spec_sheet_classes():
    """Import spec sheet generator classes - handles the hyphenated filename"""
    import sys
    import importlib.util
    
    # Get the spec sheet generator module path
    spec_file = os.path.join(os.path.dirname(__file__), '..', 'spec-sheet', 'spec-sheet-generator.py')
    
    # Load module using importlib
    spec = importlib.util.spec_from_file_location("spec_sheet_generator", spec_file)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    
    return {
        'TeamMember': module.TeamMember,
        'Team': module.Team,
        'EnhancedSpecSheetSync': module.EnhancedSpecSheetSync
    }

@pytest.fixture
def sample_team_data():
    """Sample team composition data for testing"""
    return [
        {"role": "Senior Developer", "fte": 1.0, "story_points_per_sprint": 8, "hourly_rate": 110},
        {"role": "Junior Developer", "fte": 1.0, "story_points_per_sprint": 5, "hourly_rate": 75},
        {"role": "Designer", "fte": 0.5, "story_points_per_sprint": 6, "hourly_rate": 85}
    ]

# Utility functions for tests

def assert_workbook_has_sheet(workbook_path: str, sheet_name: str):
    """Assert that workbook contains the specified sheet"""
    wb = openpyxl.load_workbook(workbook_path)
    assert sheet_name in wb.sheetnames, f"Sheet '{sheet_name}' not found in workbook"

def get_sheet_data(workbook_path: str, sheet_name: str) -> List[List]:
    """Get all data from a worksheet"""
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]
    
    data = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):  # Skip empty rows
            data.append(list(row))
    
    return data

def count_non_empty_rows(workbook_path: str, sheet_name: str) -> int:
    """Count non-empty rows in a sheet"""
    data = get_sheet_data(workbook_path, sheet_name)
    return len([row for row in data if any(cell is not None for cell in row)]) 