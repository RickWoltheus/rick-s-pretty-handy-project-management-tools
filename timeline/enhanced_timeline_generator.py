#!/usr/bin/env python3
"""
Enhanced Timeline Generator with Team Integration
Maintains original timeline behavior but uses persistent team management system:
- Reads story points from spec sheet (like original)
- Only asks for start date (like original) 
- Creates same Excel format (like original)
- Integrates with persistent team management (new feature)
"""

import sys
import os
import requests
from datetime import datetime, timedelta, date
from typing import List, Dict, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import calendar
from dataclasses import dataclass, field

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Import team management
from team.team_manager import TeamManager, TeamMember, Holiday
from utils.config import JiraConfig
from utils.jira_client import JiraClient

@dataclass
class Sprint:
    """Sprint information with dates"""
    number: int
    start_date: date
    end_date: date
    story_points: float
    name: str = ""
    team_velocity: float = 0.0

@dataclass
class HolidayInfo:
    """Holiday information for the timeline"""
    date: date
    name: str
    is_national: bool = True
    affected_members: List[str] = field(default_factory=list)  # Empty means affects everyone

class DutchHolidayAPI:
    """Fetches Dutch national holidays from open API"""
    
    @staticmethod
    def fetch_holidays(year: int) -> List[HolidayInfo]:
        """Fetch Dutch holidays for a specific year"""
        try:
            # Using Nederlandse feestdagen API
            url = f"https://date.nager.at/api/v3/PublicHolidays/{year}/NL"
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            
            holidays = []
            for holiday_data in response.json():
                holiday_date = datetime.strptime(holiday_data['date'], '%Y-%m-%d').date()
                holidays.append(HolidayInfo(
                    date=holiday_date,
                    name=holiday_data['name'],
                    is_national=True
                ))
            
            print(f"📅 Fetched {len(holidays)} Dutch national holidays for {year}")
            return holidays
            
        except Exception as e:
            print(f"⚠️  Could not fetch Dutch holidays for {year}: {e}")
            # Return some common Dutch holidays as fallback
            return DutchHolidayAPI._get_fallback_holidays(year)
    
    @staticmethod
    def _get_fallback_holidays(year: int) -> List[HolidayInfo]:
        """Fallback holidays when API is unavailable"""
        common_holidays = [
            (1, 1, "Nieuwjaarsdag"),
            (4, 27, "Koningsdag"),
            (5, 1, "Dag van de Arbeid"),
            (12, 25, "Eerste Kerstdag"),
            (12, 26, "Tweede Kerstdag"),
        ]
        
        holidays = []
        for month, day, name in common_holidays:
            try:
                holiday_date = date(year, month, day)
                holidays.append(HolidayInfo(
                    date=holiday_date,
                    name=name,
                    is_national=True
                ))
            except ValueError:
                continue  # Skip invalid dates
        
        return holidays

class EnhancedTimelineGenerator:
    """Enhanced timeline generator that loads team from persistent storage"""
    
    def __init__(self):
        self.team_manager = TeamManager()
        self.team_members: List[TeamMember] = []
        self.holidays: List[HolidayInfo] = []
        self.sprints: List[Sprint] = []
        self.project_start_date: Optional[date] = None
        self.project_end_date: Optional[date] = None
        
        # Try to connect to JIRA for integration
        try:
            self.jira_config = JiraConfig()
            self.jira_client = JiraClient(self.jira_config)
            self.jira_available = True
            print("✅ JIRA integration available")
        except:
            self.jira_available = False
            print("⚠️  JIRA not available - timeline will work without JIRA integration")
    
    def load_team_from_storage(self) -> bool:
        """Load team members from persistent storage"""
        self.team_members = self.team_manager.load_all_members()
        
        if not self.team_members:
            print("❌ No team members found in storage")
            return False
        
        print(f"✅ Loaded {len(self.team_members)} team members from storage:")
        for member in self.team_members:
            holiday_count = len(member.holidays)
            print(f"   👤 {member.name} ({member.role}) - {member.availability*100:.0f}% - {holiday_count} holidays")
        
        # Convert team member holidays to timeline holidays
        for member in self.team_members:
            for holiday in member.holidays:
                # Convert date strings to date objects
                start_date = datetime.strptime(holiday.start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(holiday.end_date, '%Y-%m-%d').date()
                
                # Add each day in the holiday range
                current_date = start_date
                while current_date <= end_date:
                    timeline_holiday = HolidayInfo(
                        date=current_date,
                        name=holiday.name,
                        is_national=holiday.is_national,
                        affected_members=[member.name] if not holiday.is_national else []
                    )
                    self.holidays.append(timeline_holiday)
                    current_date += timedelta(days=1)
        
        return True
    
    def load_dutch_holidays(self, start_year: int, end_year: int):
        """Load Dutch national holidays for the project timeline"""
        print(f"🇳🇱 Loading Dutch holidays from {start_year} to {end_year}...")
        
        for year in range(start_year, end_year + 1):
            year_holidays = DutchHolidayAPI.fetch_holidays(year)
            self.holidays.extend(year_holidays)
        
        print(f"📅 Loaded {len([h for h in self.holidays if h.is_national])} national holidays")
    
    def calculate_working_days(self, start_date: date, end_date: date, member: TeamMember) -> int:
        """Calculate working days for a team member considering holidays"""
        working_days = 0
        current_date = start_date
        
        while current_date <= end_date:
            # Check if it's a weekday (Monday = 0, Sunday = 6)
            if current_date.weekday() < 5:  # Monday to Friday
                # Check if member is on holiday
                is_on_holiday = False
                for holiday in self.holidays:
                    if (holiday.date == current_date and 
                        (holiday.is_national or member.name in holiday.affected_members)):
                        is_on_holiday = True
                        break
                
                if not is_on_holiday:
                    working_days += 1
            
            current_date += timedelta(days=1)
        
        return working_days
    
    def calculate_effective_capacity(self, member: TeamMember, start_date: date, end_date: date) -> float:
        """Calculate effective capacity for a team member"""
        working_days = self.calculate_working_days(start_date, end_date, member)
        return working_days * member.availability
    
    def generate_sprint_timeline(self, start_date: date, total_story_points: float, sprint_length_weeks: int = 2) -> List[Sprint]:
        """Generate sprint timeline based on team capacity"""
        if not self.team_members:
            print("❌ No team members available for sprint planning")
            return []
        
        # Calculate total team capacity per sprint
        sprint_days = sprint_length_weeks * 5  # Working days only
        
        # Calculate team velocity (story points per sprint)
        total_team_velocity = sum(member.story_points_per_sprint for member in self.team_members)
        
        print(f"📊 Team Analysis:")
        print(f"   Sprint length: {sprint_length_weeks} weeks ({sprint_days} working days)")
        print(f"   Team velocity: {total_team_velocity} story points per sprint")
        print(f"   Total story points needed: {total_story_points}")
        
        # Calculate number of sprints needed
        sprints_needed = max(1, round(total_story_points / total_team_velocity))
        actual_sprints_needed = total_story_points / total_team_velocity
        
        print(f"   Sprints needed: {actual_sprints_needed:.1f} (rounded to {sprints_needed})")
        
        # Generate sprints
        sprints = []
        remaining_story_points = total_story_points
        current_start_date = start_date
        
        for sprint_num in range(1, sprints_needed + 1):
            # Calculate sprint end date (skip weekends)
            sprint_end_date = current_start_date
            days_added = 0
            
            while days_added < sprint_days:
                sprint_end_date += timedelta(days=1)
                if sprint_end_date.weekday() < 5:  # Only count weekdays
                    days_added += 1
            
            # Calculate story points for this sprint
            if sprint_num == sprints_needed:
                # Last sprint gets remaining points
                sprint_story_points = remaining_story_points
            else:
                sprint_story_points = min(total_team_velocity, remaining_story_points)
            
            sprint = Sprint(
                number=sprint_num,
                start_date=current_start_date,
                end_date=sprint_end_date,
                story_points=sprint_story_points,
                name=f"Sprint {sprint_num}",
                team_velocity=total_team_velocity
            )
            
            sprints.append(sprint)
            remaining_story_points -= sprint_story_points
            
            # Next sprint starts the Monday after this one ends
            current_start_date = sprint_end_date + timedelta(days=1)
            while current_start_date.weekday() != 0:  # Find next Monday
                current_start_date += timedelta(days=1)
        
        self.sprints = sprints
        self.project_start_date = start_date
        self.project_end_date = sprints[-1].end_date if sprints else start_date
        
        print(f"\n📅 Generated {len(sprints)} sprints:")
        for sprint in sprints:
            print(f"   Sprint {sprint.number}: {sprint.start_date} to {sprint.end_date} ({sprint.story_points:.1f} SP)")
        
        return sprints
    
    def load_story_points_from_spec_sheet(self, spec_sheet_path: str = "spec-sheet.xlsx") -> float:
        """Load total story points from existing spec sheet (same as original)"""
        try:
            if not os.path.exists(spec_sheet_path):
                print(f"⚠️  Spec sheet not found: {spec_sheet_path}")
                return 0
            
            wb = openpyxl.load_workbook(spec_sheet_path, data_only=True)
            print(f"📊 Available sheets in {spec_sheet_path}: {wb.sheetnames}")
            
            # Try to find story points in various common sheet names
            sheet_names = ['Scope (Quantity)', 'Scope', 'Timeline', 'Summary', 'Sheet1']
            total_story_points = 0
            found_sheet = None
            
            for sheet_name in sheet_names:
                if sheet_name in wb.sheetnames:
                    found_sheet = sheet_name
                    break
            
            # If no common sheet names found, try the first sheet
            if not found_sheet and wb.sheetnames:
                found_sheet = wb.sheetnames[0]
                print(f"📊 Using first available sheet: {found_sheet}")
            
            if found_sheet:
                ws = wb[found_sheet]
                print(f"📊 Analyzing sheet: {found_sheet}")
                print(f"📊 Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
                
                # Look for story points columns
                story_point_cols = []
                header_info = {}
                
                print("\n🔍 Searching for story point columns...")
                for row_num in range(1, min(10, ws.max_row + 1)):  # Check first 10 rows for headers
                    for col_num in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.value and isinstance(cell.value, str):
                            cell_value_lower = cell.value.lower()
                            # Broader keyword search
                            keywords = ['story point', 'sp', 'points', 'story pts', 'storypoint', 'story_point']
                            for keyword in keywords:
                                if keyword in cell_value_lower:
                                    story_point_cols.append(col_num)
                                    header_info[col_num] = f"Row {row_num}: '{cell.value}'"
                                    print(f"   ✅ Found potential story point column: {col_num} (Row {row_num}: '{cell.value}')")
                                    break
                
                # Sum up story points from found columns
                for col in story_point_cols:
                    col_total = 0
                    print(f"\n📊 Processing column {col} ({header_info.get(col, 'Unknown header')}):")
                    
                    for row_num in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row_num, column=col).value
                        
                        # Handle different story point formats
                        story_point_value = None
                        
                        # Check if it's already a number
                        if isinstance(cell_value, (int, float)) and cell_value > 0:
                            story_point_value = cell_value
                        
                        # Check if it's text containing "Story Points: X"
                        elif isinstance(cell_value, str):
                            import re
                            # Look for patterns like "Story Points: 5" or "SP: 3" etc.
                            patterns = [
                                r'story\s*points?\s*[:\-]?\s*(\d+(?:\.\d+)?)',
                                r'sp\s*[:\-]?\s*(\d+(?:\.\d+)?)',
                                r'points?\s*[:\-]?\s*(\d+(?:\.\d+)?)'
                            ]
                            
                            for pattern in patterns:
                                match = re.search(pattern, cell_value.lower())
                                if match:
                                    try:
                                        story_point_value = float(match.group(1))
                                        break
                                    except ValueError:
                                        continue
                        
                        if story_point_value is not None and story_point_value > 0:
                            col_total += story_point_value
                            total_story_points += story_point_value
                            print(f"   Row {row_num}: {story_point_value} SP (from '{cell_value}')")
                    
                    print(f"   Column {col} total: {col_total} SP")
                
                print(f"\n📋 Total story points found: {total_story_points}")
            
            if total_story_points > 0:
                print(f"✅ Successfully loaded {total_story_points} story points from {spec_sheet_path}")
                return total_story_points
            else:
                print("⚠️  No story points found in spec sheet")
                return 0
            
        except Exception as e:
            print(f"❌ Error reading spec sheet: {e}")
            return 0
    
    def save_timeline_workbook(self, filename: str = None):
        """Save the complete timeline workbook (same format as original)"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"project_timeline_{timestamp}.xlsx"
        
        print(f"📊 Creating timeline workbook: {filename}")
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create all sheets
        self._create_team_members_sheet(wb)
        self._create_holidays_sheet(wb)
        self._create_sprint_timeline_sheet(wb)
        self._create_calendar_view_sheet(wb)
        
        # Save workbook
        wb.save(filename)
        print(f"✅ Timeline workbook saved: {filename}")
        
        return filename
    
    def _create_team_members_sheet(self, wb: openpyxl.Workbook):
        """Create team members sheet with enhanced information"""
        ws = wb.create_sheet("Team Members")
        
        # Headers
        headers = ["Name", "Role", "Availability", "SP/Sprint", "Hourly Rate", "Email", "Holidays"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Team member data
        for row, member in enumerate(self.team_members, 2):
            ws.cell(row=row, column=1, value=member.name)
            ws.cell(row=row, column=2, value=member.role)
            ws.cell(row=row, column=3, value=f"{member.availability*100:.0f}%")
            ws.cell(row=row, column=4, value=member.story_points_per_sprint)
            ws.cell(row=row, column=5, value=f"€{member.hourly_rate:.2f}")
            ws.cell(row=row, column=6, value=member.email or "")
            ws.cell(row=row, column=7, value=len(member.holidays))
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_holidays_sheet(self, wb: openpyxl.Workbook):
        """Create holidays sheet"""
        ws = wb.create_sheet("Holidays")
        
        # Headers
        headers = ["Date", "Name", "Type", "Affected Members"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Holiday data
        sorted_holidays = sorted(self.holidays, key=lambda h: h.date)
        for row, holiday in enumerate(sorted_holidays, 2):
            ws.cell(row=row, column=1, value=holiday.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=holiday.name)
            ws.cell(row=row, column=3, value="National" if holiday.is_national else "Personal")
            ws.cell(row=row, column=4, value=", ".join(holiday.affected_members) if holiday.affected_members else "Everyone")
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_sprint_timeline_sheet(self, wb: openpyxl.Workbook):
        """Create sprint timeline sheet (main timeline view)"""
        ws = wb.create_sheet("Sprint Timeline")
        
        # Headers
        headers = ["Sprint", "Start Date", "End Date", "Story Points", "Team Velocity", "Duration (Days)", "Capacity Utilization"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Sprint data
        for row, sprint in enumerate(self.sprints, 2):
            duration = (sprint.end_date - sprint.start_date).days + 1
            utilization = (sprint.story_points / sprint.team_velocity * 100) if sprint.team_velocity > 0 else 0
            
            ws.cell(row=row, column=1, value=f"Sprint {sprint.number}")
            ws.cell(row=row, column=2, value=sprint.start_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=3, value=sprint.end_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=4, value=sprint.story_points)
            ws.cell(row=row, column=5, value=sprint.team_velocity)
            ws.cell(row=row, column=6, value=duration)
            ws.cell(row=row, column=7, value=f"{utilization:.1f}%")
        
        # Summary row
        if self.sprints:
            summary_row = len(self.sprints) + 3
            ws.cell(row=summary_row, column=1, value="TOTAL")
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            ws.cell(row=summary_row, column=4, value=sum(s.story_points for s in self.sprints))
            ws.cell(row=summary_row, column=6, value=(self.project_end_date - self.project_start_date).days + 1)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_calendar_view_sheet(self, wb: openpyxl.Workbook):
        """Create a calendar view showing team availability and project timeline"""
        ws = wb.create_sheet("Calendar View")
        
        if not self.project_start_date or not self.project_end_date:
            ws.cell(row=1, column=1, value="No project timeline generated yet")
            return
        
        # Title
        ws.cell(row=1, column=1, value=f"Project Calendar: {self.project_start_date} to {self.project_end_date}")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        
        # Legend
        ws.cell(row=2, column=1, value="Legend:")
        ws.cell(row=2, column=2, value="Sprint Days")
        ws.cell(row=2, column=2).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        ws.cell(row=2, column=3, value="Holidays")
        ws.cell(row=2, column=3).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        ws.cell(row=2, column=4, value="Weekends")
        ws.cell(row=2, column=4).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Generate monthly calendar views
        current_month = self.project_start_date.replace(day=1)
        end_month = self.project_end_date.replace(day=1)
        row_offset = 5
        
        while current_month <= end_month:
            row_offset = self._create_monthly_calendar(ws, current_month, row_offset)
            # Move to next month
            if current_month.month == 12:
                current_month = current_month.replace(year=current_month.year + 1, month=1)
            else:
                current_month = current_month.replace(month=current_month.month + 1)
    
    def _create_monthly_calendar(self, ws: openpyxl.Workbook, month_date: date, start_row: int) -> int:
        """Create a single month calendar view"""
        month_name = month_date.strftime("%B %Y")
        
        # Month header
        ws.cell(row=start_row, column=1, value=month_name)
        ws.cell(row=start_row, column=1).font = Font(size=12, bold=True)
        
        # Day headers
        day_headers = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for col, day in enumerate(day_headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=day)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Get calendar for this month
        cal = calendar.monthcalendar(month_date.year, month_date.month)
        
        # Fill in the calendar
        for week_num, week in enumerate(cal):
            row = start_row + 2 + week_num
            for day_num, day in enumerate(week):
                col = day_num + 1
                
                if day == 0:
                    continue  # Empty cell for days not in this month
                
                current_date = date(month_date.year, month_date.month, day)
                cell = ws.cell(row=row, column=col, value=day)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Check if this date is special
                cell_fill = None
                
                # Check for holidays
                for holiday in self.holidays:
                    if holiday.date == current_date:
                        cell_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        break
                
                # Check for sprint dates
                for sprint in self.sprints:
                    if sprint.start_date <= current_date <= sprint.end_date:
                        if cell_fill is None:  # Don't override holiday colors
                            cell_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        break
                
                # Check for weekends
                if current_date.weekday() >= 5:  # Saturday = 5, Sunday = 6
                    if cell_fill is None:
                        cell_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                
                if cell_fill:
                    cell.fill = cell_fill
                
                # Add border
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Set column widths
        for col in range(1, 8):
            column_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[column_letter].width = 6
        
        return start_row + len(cal) + 4  # Return next available row
    
    def run_interactive(self):
        """Run the enhanced timeline generator interactively"""
        print("\n📅 ENHANCED TIMELINE GENERATOR")
        print("=" * 50)
        
        # Load team from storage
        if not self.load_team_from_storage():
            print("❌ Cannot continue without team members")
            print("💡 Please create team members first using Team Management")
            return
        
        # Load story points from spec sheet
        print("\n📊 Loading story points from spec sheet...")
        total_story_points = self.load_story_points_from_spec_sheet()
        
        if total_story_points <= 0:
            print("❌ No story points found in spec sheet")
            print("💡 Please generate a spec sheet first or ensure it contains story points")
            return
        
        # Get project start date
        print(f"\n📅 Project Planning ({total_story_points} story points)")
        try:
            start_date_str = input("Project start date (YYYY-MM-DD): ").strip()
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            print("❌ Invalid date format")
            return
        
        # Load Dutch holidays
        start_year = start_date.year
        end_year = start_year + 1  # Load holidays for current and next year
        self.load_dutch_holidays(start_year, end_year)
        
        # Generate sprint timeline
        print(f"\n🔄 Generating sprint timeline from {start_date}...")
        self.generate_sprint_timeline(start_date, total_story_points)
        
        # Save timeline workbook
        filename = self.save_timeline_workbook()
        
        print(f"\n✅ Enhanced timeline generation complete!")
        print(f"📊 Timeline saved to: {filename}")
        print(f"🎯 Project duration: {self.project_start_date} to {self.project_end_date}")
        print(f"📈 Total sprints: {len(self.sprints)}")
        print(f"👥 Team members: {len(self.team_members)}")

def main():
    """Enhanced timeline generator main function"""
    generator = EnhancedTimelineGenerator()
    
    while True:
        print("\n" + "=" * 50)
        print("📅 ENHANCED TIMELINE GENERATOR")
        print("=" * 50)
        print("1. 📊 Generate Project Timeline")
        print("2. 👥 View Team Members")
        print("3. 🏖️  Manage Team (Create/Edit Members)")
        print("0. 🚪 Back to Main Menu")
        print("=" * 50)
        
        choice = input("Select option (0-3): ").strip()
        
        if choice == '0':
            break
        elif choice == '1':
            generator.run_interactive()
        elif choice == '2':
            generator.team_manager.list_all_members()
        elif choice == '3':
            from team.team_manager import main as team_main
            team_main()
        else:
            print("❌ Invalid choice")

if __name__ == "__main__":
    main() 