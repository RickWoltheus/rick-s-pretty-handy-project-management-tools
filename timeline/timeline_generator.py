#!/usr/bin/env python3
"""
Timeline and Calendar Management Tool
Generates project timelines with team availability, holidays, and calendar views
Integrates with the existing spec sheet system for comprehensive project planning
"""

import sys
import os
import requests
from datetime import datetime, timedelta, date
from typing import List, Dict, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import calendar
from dataclasses import dataclass, field

# Add project root to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from utils.config import JiraConfig
from utils.jira_client import JiraClient

@dataclass
class TeamMemberInfo:
    """Enhanced team member information for timeline planning"""
    name: str
    role: str
    fte: float  # Full Time Equivalent (0.0 to 1.0)
    story_points_per_sprint: float
    hourly_rate: float
    start_date: Optional[date] = None
    end_date: Optional[date] = None  # For temporary team members
    seniority_level: str = "Mid"  # Junior, Mid, Senior, Lead
    skills: List[str] = field(default_factory=list)
    custom_holidays: List[date] = field(default_factory=list)
    working_days: List[int] = field(default_factory=lambda: [0, 1, 2, 3, 4])  # Mon-Fri default
    
    def __post_init__(self):
        if self.start_date is None:
            self.start_date = date.today()

@dataclass
class HolidayInfo:
    """Holiday information"""
    date: date
    name: str
    is_national: bool = True
    affected_members: List[str] = field(default_factory=list)  # Empty means affects everyone

@dataclass
class Sprint:
    """Sprint information with dates"""
    number: int
    start_date: date
    end_date: date
    story_points: float
    name: str = ""
    team_velocity: float = 0.0

class DutchHolidayAPI:
    """Fetches Dutch national holidays from open API"""
    
    @staticmethod
    def fetch_holidays(year: int) -> List[HolidayInfo]:
        """Fetch Dutch holidays for a specific year"""
        try:
            # Using Nederlandse feestdagen API
            url = f"https://date.nager.at/api/v3/PublicHolidays/{year}/NL"
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            
            holidays = []
            for holiday_data in response.json():
                holiday_date = datetime.strptime(holiday_data['date'], '%Y-%m-%d').date()
                holidays.append(HolidayInfo(
                    date=holiday_date,
                    name=holiday_data['name'],
                    is_national=True
                ))
            
            print(f"📅 Fetched {len(holidays)} Dutch national holidays for {year}")
            return holidays
            
        except Exception as e:
            print(f"⚠️  Could not fetch Dutch holidays for {year}: {e}")
            # Return some common Dutch holidays as fallback
            return DutchHolidayAPI._get_fallback_holidays(year)
    
    @staticmethod
    def _get_fallback_holidays(year: int) -> List[HolidayInfo]:
        """Fallback holidays when API is unavailable"""
        common_holidays = [
            (1, 1, "Nieuwjaarsdag"),
            (4, 27, "Koningsdag"),
            (5, 1, "Dag van de Arbeid"),
            (12, 25, "Eerste Kerstdag"),
            (12, 26, "Tweede Kerstdag"),
        ]
        
        holidays = []
        for month, day, name in common_holidays:
            try:
                holiday_date = date(year, month, day)
                holidays.append(HolidayInfo(
                    date=holiday_date,
                    name=name,
                    is_national=True
                ))
            except ValueError:
                continue  # Skip invalid dates
        
        return holidays

class TimelineGenerator:
    """Main timeline and calendar generator"""
    
    def __init__(self):
        self.team_members: List[TeamMemberInfo] = []
        self.holidays: List[HolidayInfo] = []
        self.sprints: List[Sprint] = []
        self.project_start_date: Optional[date] = None
        self.project_end_date: Optional[date] = None
        
        # Try to connect to JIRA for integration
        try:
            self.jira_config = JiraConfig()
            self.jira_client = JiraClient(self.jira_config)
            self.jira_available = True
            print("✅ JIRA integration available")
        except:
            self.jira_available = False
            print("⚠️  JIRA not available - timeline will work without JIRA integration")

    def add_team_member(self, member: TeamMemberInfo):
        """Add a team member to the timeline"""
        self.team_members.append(member)
        print(f"👤 Added team member: {member.name} ({member.role}, {member.fte} FTE)")

    def load_dutch_holidays(self, start_year: int, end_year: int):
        """Load Dutch national holidays for the project timeline"""
        print(f"🇳🇱 Loading Dutch holidays from {start_year} to {end_year}...")
        
        for year in range(start_year, end_year + 1):
            year_holidays = DutchHolidayAPI.fetch_holidays(year)
            self.holidays.extend(year_holidays)
        
        print(f"📅 Loaded {len(self.holidays)} total holidays")

    def add_custom_holiday(self, holiday: HolidayInfo):
        """Add a custom holiday or team member specific holiday"""
        self.holidays.append(holiday)
        affected = "everyone" if not holiday.affected_members else ", ".join(holiday.affected_members)
        print(f"🏖️  Added custom holiday: {holiday.name} on {holiday.date} (affects: {affected})")

    def add_holiday_range(self, start_date: date, end_date: date, name: str, is_national: bool = False, affected_members: List[str] = None):
        """Add a range of holidays (e.g., vacation period, company shutdown)"""
        if affected_members is None:
            affected_members = []
        
        if start_date > end_date:
            print("❌ Error: Start date must be before or equal to end date")
            return
        
        holidays_added = 0
        current_date = start_date
        
        while current_date <= end_date:
            # Create a holiday for each day in the range
            day_name = f"{name} - Day {holidays_added + 1}" if (end_date - start_date).days > 0 else name
            
            holiday = HolidayInfo(
                date=current_date,
                name=day_name,
                is_national=is_national,
                affected_members=affected_members.copy()
            )
            
            self.holidays.append(holiday)
            holidays_added += 1
            current_date += timedelta(days=1)
        
        affected = "everyone" if not affected_members else ", ".join(affected_members)
        print(f"🏖️  Added holiday range: {name} from {start_date} to {end_date} ({holidays_added} days, affects: {affected})")
        
        return holidays_added

    def add_workdays_holiday_range(self, start_date: date, end_date: date, name: str, is_national: bool = False, affected_members: List[str] = None):
        """Add a range of holidays but only for working days (Mon-Fri)"""
        if affected_members is None:
            affected_members = []
        
        if start_date > end_date:
            print("❌ Error: Start date must be before or equal to end date")
            return
        
        holidays_added = 0
        current_date = start_date
        
        while current_date <= end_date:
            # Only add holidays for Monday through Friday (0-4)
            if current_date.weekday() < 5:  # Monday=0, Friday=4
                day_name = f"{name} - Day {holidays_added + 1}" if (end_date - start_date).days > 4 else name
                
                holiday = HolidayInfo(
                    date=current_date,
                    name=day_name,
                    is_national=is_national,
                    affected_members=affected_members.copy()
                )
                
                self.holidays.append(holiday)
                holidays_added += 1
            
            current_date += timedelta(days=1)
        
        affected = "everyone" if not affected_members else ", ".join(affected_members)
        print(f"🏖️  Added workdays holiday range: {name} from {start_date} to {end_date} ({holidays_added} working days, affects: {affected})")
        
        return holidays_added

    def calculate_working_days(self, start_date: date, end_date: date, member: TeamMemberInfo) -> int:
        """Calculate working days for a team member between two dates"""
        working_days = 0
        current_date = start_date
        
        while current_date <= end_date:
            # Check if it's a working day for this member
            if current_date.weekday() in member.working_days:
                # Check if it's not a holiday
                is_holiday = False
                for holiday in self.holidays:
                    if holiday.date == current_date:
                        # Check if this holiday affects this member
                        if not holiday.affected_members or member.name in holiday.affected_members:
                            is_holiday = True
                            break
                
                # Check custom holidays for this member
                if current_date in member.custom_holidays:
                    is_holiday = True
                
                if not is_holiday:
                    working_days += 1
            
            current_date += timedelta(days=1)
        
        return working_days

    def calculate_effective_capacity(self, member: TeamMemberInfo, start_date: date, end_date: date) -> float:
        """Calculate effective story point capacity for a member over a period"""
        total_days = (end_date - start_date).days + 1
        working_days = self.calculate_working_days(start_date, end_date, member)
        
        # Calculate capacity considering FTE and actual working days
        if total_days > 0:
            # Assume 2-week sprints as baseline
            baseline_sprint_days = 10  # 2 weeks of working days
            effective_capacity = member.story_points_per_sprint * (working_days / baseline_sprint_days) * member.fte
            return max(0, effective_capacity)
        
        return 0

    def generate_sprint_timeline(self, start_date: date, total_story_points: float, sprint_length_weeks: int = 2) -> List[Sprint]:
        """Generate sprint timeline based on team capacity and story points"""
        if not self.team_members:
            print("⚠️  No team members added. Cannot generate timeline.")
            return []
        
        print(f"\n📊 Generating sprint timeline from {start_date}")
        print(f"📋 Total story points to deliver: {total_story_points}")
        print(f"⏱️  Sprint length: {sprint_length_weeks} weeks")
        
        sprints = []
        current_date = start_date
        remaining_story_points = total_story_points
        sprint_number = 1
        
        sprint_length_days = sprint_length_weeks * 7
        
        while remaining_story_points > 0:
            sprint_start = current_date
            sprint_end = current_date + timedelta(days=sprint_length_days - 1)
            
            # Calculate team capacity for this sprint
            team_capacity = 0
            for member in self.team_members:
                member_capacity = self.calculate_effective_capacity(member, sprint_start, sprint_end)
                team_capacity += member_capacity
            
            # Determine story points for this sprint
            sprint_story_points = min(remaining_story_points, team_capacity)
            
            sprint = Sprint(
                number=sprint_number,
                start_date=sprint_start,
                end_date=sprint_end,
                story_points=sprint_story_points,
                name=f"Sprint {sprint_number}",
                team_velocity=team_capacity
            )
            
            sprints.append(sprint)
            remaining_story_points -= sprint_story_points
            
            print(f"   Sprint {sprint_number}: {sprint_start} to {sprint_end} - {sprint_story_points:.1f} SP (capacity: {team_capacity:.1f})")
            
            current_date = sprint_end + timedelta(days=1)
            sprint_number += 1
            
            # Safety break to prevent infinite loops
            if sprint_number > 50:
                print("⚠️  Reached maximum sprint limit (50). Timeline may be incomplete.")
                break
        
        self.sprints = sprints
        self.project_start_date = start_date
        self.project_end_date = sprints[-1].end_date if sprints else start_date
        
        print(f"\n✅ Generated {len(sprints)} sprints")
        print(f"📅 Project timeline: {self.project_start_date} to {self.project_end_date}")
        
        return sprints

    def save_timeline_workbook(self, filename: str = None):
        """Save the timeline workbook"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"project_timeline_{timestamp}.xlsx"
        
        wb = self.create_team_sheet()
        wb.save(filename)
        print(f"💾 Timeline saved: {filename}")
        return filename

    def create_team_sheet(self) -> openpyxl.Workbook:
        """Create an Excel workbook with team information and timeline"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Team Members sheet
        self._create_team_members_sheet(wb)
        
        # Create Holidays sheet
        self._create_holidays_sheet(wb)
        
        # Create Sprint Timeline sheet
        self._create_sprint_timeline_sheet(wb)
        
        # Create Calendar View sheet
        self._create_calendar_view_sheet(wb)
        
        return wb

    def _create_team_members_sheet(self, wb: openpyxl.Workbook):
        """Create the team members information sheet"""
        ws = wb.create_sheet("Team Members")
        
        # Headers
        headers = [
            "Name", "Role", "FTE", "SP/Sprint", "Hourly Rate", "Start Date", "End Date",
            "Seniority", "Skills", "Working Days", "Custom Holidays"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Team member data
        for row, member in enumerate(self.team_members, 2):
            ws.cell(row=row, column=1, value=member.name)
            ws.cell(row=row, column=2, value=member.role)
            ws.cell(row=row, column=3, value=member.fte)
            ws.cell(row=row, column=4, value=member.story_points_per_sprint)
            ws.cell(row=row, column=5, value=member.hourly_rate)
            ws.cell(row=row, column=6, value=member.start_date)
            ws.cell(row=row, column=7, value=member.end_date)
            ws.cell(row=row, column=8, value=member.seniority_level)
            ws.cell(row=row, column=9, value=", ".join(member.skills))
            working_days_names = [calendar.day_name[d] for d in member.working_days]
            ws.cell(row=row, column=10, value=", ".join(working_days_names))
            ws.cell(row=row, column=11, value=", ".join([str(h) for h in member.custom_holidays]))
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_holidays_sheet(self, wb: openpyxl.Workbook):
        """Create the holidays management sheet"""
        ws = wb.create_sheet("Holidays")
        
        # Headers
        headers = ["Date", "Holiday Name", "Type", "Affected Members"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Holiday data
        sorted_holidays = sorted(self.holidays, key=lambda h: h.date)
        for row, holiday in enumerate(sorted_holidays, 2):
            ws.cell(row=row, column=1, value=holiday.date)
            ws.cell(row=row, column=2, value=holiday.name)
            ws.cell(row=row, column=3, value="National" if holiday.is_national else "Custom")
            affected = "Everyone" if not holiday.affected_members else ", ".join(holiday.affected_members)
            ws.cell(row=row, column=4, value=affected)
            
            # Color coding
            if holiday.is_national:
                fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFF2E8", end_color="FFF2E8", fill_type="solid")
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_sprint_timeline_sheet(self, wb: openpyxl.Workbook):
        """Create the sprint timeline sheet"""
        ws = wb.create_sheet("Sprint Timeline")
        
        # Headers
        headers = [
            "Sprint", "Start Date", "End Date", "Duration (Days)", "Story Points", 
            "Team Velocity", "Working Days", "Efficiency %"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Sprint data
        for row, sprint in enumerate(self.sprints, 2):
            duration = (sprint.end_date - sprint.start_date).days + 1
            
            # Calculate total working days for the team
            total_working_days = 0
            for member in self.team_members:
                member_working_days = self.calculate_working_days(sprint.start_date, sprint.end_date, member)
                total_working_days += member_working_days * member.fte
            
            efficiency = (sprint.story_points / sprint.team_velocity * 100) if sprint.team_velocity > 0 else 0
            
            ws.cell(row=row, column=1, value=sprint.name)
            ws.cell(row=row, column=2, value=sprint.start_date)
            ws.cell(row=row, column=3, value=sprint.end_date)
            ws.cell(row=row, column=4, value=duration)
            ws.cell(row=row, column=5, value=round(sprint.story_points, 1))
            ws.cell(row=row, column=6, value=round(sprint.team_velocity, 1))
            ws.cell(row=row, column=7, value=round(total_working_days, 1))
            ws.cell(row=row, column=8, value=f"{efficiency:.1f}%")
        
        # Add summary row
        if self.sprints:
            summary_row = len(self.sprints) + 3
            ws.cell(row=summary_row, column=1, value="TOTAL")
            ws.cell(row=summary_row, column=5, value=sum(s.story_points for s in self.sprints))
            ws.cell(row=summary_row, column=6, value=sum(s.team_velocity for s in self.sprints))
            
            # Style summary row
            for col in range(1, 9):
                cell = ws.cell(row=summary_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_calendar_view_sheet(self, wb: openpyxl.Workbook):
        """Create a calendar view showing team availability and project timeline"""
        ws = wb.create_sheet("Calendar View")
        
        if not self.project_start_date or not self.project_end_date:
            ws.cell(row=1, column=1, value="No project timeline generated yet")
            return
        
        # Title
        ws.cell(row=1, column=1, value=f"Project Calendar: {self.project_start_date} to {self.project_end_date}")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        
        # Legend
        ws.cell(row=2, column=1, value="Legend:")
        ws.cell(row=2, column=2, value="Sprint Days")
        ws.cell(row=2, column=2).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        ws.cell(row=2, column=3, value="Holidays")
        ws.cell(row=2, column=3).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        ws.cell(row=2, column=4, value="Weekends")
        ws.cell(row=2, column=4).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Generate monthly calendar views
        current_month = self.project_start_date.replace(day=1)
        end_month = self.project_end_date.replace(day=1)
        row_offset = 5
        
        while current_month <= end_month:
            row_offset = self._create_monthly_calendar(ws, current_month, row_offset)
            # Move to next month
            if current_month.month == 12:
                current_month = current_month.replace(year=current_month.year + 1, month=1)
            else:
                current_month = current_month.replace(month=current_month.month + 1)

    def _create_monthly_calendar(self, ws: openpyxl.Workbook, month_date: date, start_row: int) -> int:
        """Create a single month calendar view"""
        month_name = month_date.strftime("%B %Y")
        
        # Month header
        ws.cell(row=start_row, column=1, value=month_name)
        ws.cell(row=start_row, column=1).font = Font(size=12, bold=True)
        
        # Day headers
        day_headers = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for col, day in enumerate(day_headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=day)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Get calendar for this month
        cal = calendar.monthcalendar(month_date.year, month_date.month)
        
        # Fill in the calendar
        for week_num, week in enumerate(cal):
            row = start_row + 2 + week_num
            for day_num, day in enumerate(week):
                col = day_num + 1
                
                if day == 0:
                    continue  # Empty cell for days not in this month
                
                current_date = date(month_date.year, month_date.month, day)
                cell = ws.cell(row=row, column=col, value=day)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Check if this date is special
                cell_fill = None
                
                # Check for holidays
                for holiday in self.holidays:
                    if holiday.date == current_date:
                        cell_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        break
                
                # Check for sprint dates
                for sprint in self.sprints:
                    if sprint.start_date <= current_date <= sprint.end_date:
                        if cell_fill is None:  # Don't override holiday colors
                            cell_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        break
                
                # Check for weekends
                if current_date.weekday() >= 5:  # Saturday = 5, Sunday = 6
                    if cell_fill is None:
                        cell_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                
                if cell_fill:
                    cell.fill = cell_fill
                
                # Add border
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Set column widths
        for col in range(1, 8):
            column_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[column_letter].width = 6
        
        return start_row + len(cal) + 4  # Return next available row

    def load_story_points_from_spec_sheet(self, spec_sheet_path: str = "spec-sheet.xlsx") -> float:
        """Load total story points from existing spec sheet with detailed debugging"""
        try:
            if not os.path.exists(spec_sheet_path):
                print(f"⚠️  Spec sheet not found: {spec_sheet_path}")
                return 0
            
            wb = openpyxl.load_workbook(spec_sheet_path, data_only=True)
            print(f"📊 Available sheets in {spec_sheet_path}: {wb.sheetnames}")
            
            # Try to find story points in various common sheet names
            sheet_names = ['Scope (Quantity)', 'Scope', 'Timeline', 'Summary', 'Sheet1']
            total_story_points = 0
            found_sheet = None
            
            for sheet_name in sheet_names:
                if sheet_name in wb.sheetnames:
                    found_sheet = sheet_name
                    break
            
            # If no common sheet names found, try the first sheet
            if not found_sheet and wb.sheetnames:
                found_sheet = wb.sheetnames[0]
                print(f"📊 Using first available sheet: {found_sheet}")
            
            if found_sheet:
                ws = wb[found_sheet]
                print(f"📊 Analyzing sheet: {found_sheet}")
                print(f"📊 Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
                
                # Show first few rows for debugging
                print("🔍 First 5 rows for debugging:")
                for row_num in range(1, min(6, ws.max_row + 1)):
                    row_values = []
                    for col_num in range(1, min(10, ws.max_column + 1)):  # Show first 10 columns
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        if cell_value is not None:
                            row_values.append(f"Col{col_num}: {cell_value}")
                    if row_values:
                        print(f"   Row {row_num}: {' | '.join(row_values)}")
                
                # Look for story points columns with broader search
                story_point_cols = []
                header_info = {}
                
                print("\n🔍 Searching for story point columns...")
                for row_num in range(1, min(10, ws.max_row + 1)):  # Check first 10 rows for headers
                    for col_num in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.value and isinstance(cell.value, str):
                            cell_value_lower = cell.value.lower()
                            # Broader keyword search
                            keywords = ['story point', 'sp', 'points', 'story pts', 'storypoint', 'story_point']
                            for keyword in keywords:
                                if keyword in cell_value_lower:
                                    story_point_cols.append(col_num)
                                    header_info[col_num] = f"Row {row_num}: '{cell.value}'"
                                    print(f"   ✅ Found potential story point column: {col_num} (Row {row_num}: '{cell.value}')")
                                    break
                
                if not story_point_cols:
                    print("⚠️  No story point columns found with keywords")
                    
                # Also look for any numeric columns that might contain story points
                print("\n🔍 Looking for columns with numeric values (potential story points):")
                numeric_cols = {}
                for col_num in range(1, ws.max_column + 1):
                    col_total = 0
                    value_count = 0
                    header_name = "Unknown"
                    
                    # Get header name
                    for row_num in range(1, min(6, ws.max_row + 1)):
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                            header_name = cell.value[:50]  # First 50 chars
                            break
                    
                    # Count numeric values
                    for row_num in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        if isinstance(cell_value, (int, float)) and cell_value > 0:
                            col_total += cell_value
                            value_count += 1
                    
                    if value_count > 0:
                        numeric_cols[col_num] = {
                            'header': header_name,
                            'total': col_total,
                            'count': value_count
                        }
                        print(f"   Col {col_num} ('{header_name}'): {value_count} numeric values, total = {col_total}")
                
                # If no story point columns found but we have numeric columns, suggest them
                if not story_point_cols and numeric_cols:
                    print("\n💡 Consider if any of these numeric columns contain story points:")
                    for col_num, info in numeric_cols.items():
                        print(f"   Column {col_num}: {info['count']} values totaling {info['total']}")
                
                # Also show all headers for manual inspection
                if not story_point_cols:
                    print("\n🔍 All text headers for manual inspection:")
                    for row_num in range(1, min(10, ws.max_row + 1)):
                        for col_num in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                                print(f"   Row {row_num}, Col {col_num}: '{cell.value[:100]}{'...' if len(str(cell.value)) > 100 else ''}'")
                
                # Sum up story points from found columns
                for col in story_point_cols:
                    col_total = 0
                    print(f"\n📊 Processing column {col} ({header_info.get(col, 'Unknown header')}):")
                    
                    for row_num in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row_num, column=col).value
                        
                        # Handle different story point formats
                        story_point_value = None
                        
                        # Check if it's already a number
                        if isinstance(cell_value, (int, float)) and cell_value > 0:
                            story_point_value = cell_value
                        
                        # Check if it's text containing "Story Points: X"
                        elif isinstance(cell_value, str):
                            import re
                            # Look for patterns like "Story Points: 5" or "SP: 3" etc.
                            patterns = [
                                r'story\s*points?\s*[:\-]?\s*(\d+(?:\.\d+)?)',
                                r'sp\s*[:\-]?\s*(\d+(?:\.\d+)?)',
                                r'points?\s*[:\-]?\s*(\d+(?:\.\d+)?)'
                            ]
                            
                            for pattern in patterns:
                                match = re.search(pattern, cell_value.lower())
                                if match:
                                    try:
                                        story_point_value = float(match.group(1))
                                        break
                                    except ValueError:
                                        continue
                        
                        if story_point_value is not None:
                            if story_point_value > 0:
                                col_total += story_point_value
                                total_story_points += story_point_value
                                print(f"   Row {row_num}: {story_point_value} SP (from '{cell_value}')")
                            else:
                                print(f"   Row {row_num}: {story_point_value} SP (from '{cell_value}') - Zero value")
                        elif cell_value and cell_value != 0:
                            print(f"   Row {row_num}: '{cell_value}' (could not extract story points)")
                    
                    print(f"   Column {col} total: {col_total} SP")
                
                print(f"\n📋 Total story points found: {total_story_points}")
            
            if total_story_points > 0:
                print(f"✅ Successfully loaded {total_story_points} story points from {spec_sheet_path}")
                return total_story_points
            else:
                print("⚠️  No story points found in spec sheet")
                return 0
            
        except Exception as e:
            print(f"❌ Error reading spec sheet: {e}")
            import traceback
            traceback.print_exc()
            return 0

    def load_story_points_from_jira(self, version_filter: str = None) -> float:
        """Load total story points from JIRA for a specific version/release"""
        if not self.jira_available:
            print("⚠️  JIRA not available for story point loading")
            return 0
        
        try:
            print("📊 Loading story points from JIRA...")
            
            # Get epics for the version
            epics = self.jira_client.get_epics(version_filter)
            total_story_points = 0
            
            if not epics:
                print("⚠️  No epics found for the specified criteria")
                return 0
            
            print(f"📋 Found {len(epics)} epics")
            
            # Get all stories for each epic
            for epic in epics:
                epic_key = epic['key']
                epic_summary = epic['fields']['summary']
                print(f"   📖 Processing epic: {epic_key} - {epic_summary}")
                
                # Get stories for this epic
                stories = self.jira_client.get_stories_for_epic(epic_key)
                epic_story_points = 0
                
                for story in stories:
                    story_points = self.jira_client.get_story_points(story)
                    if story_points:
                        epic_story_points += story_points
                        total_story_points += story_points
                
                print(f"      ✅ {len(stories)} stories, {epic_story_points} story points")
            
            # Also check for issues directly assigned to the version (not in epics)
            try:
                direct_issues = self.jira_client.get_issues_for_version(version_filter, ['Story', 'Bug', 'Task'])
                direct_story_points = 0
                
                for issue in direct_issues:
                    story_points = self.jira_client.get_story_points(issue)
                    if story_points:
                        direct_story_points += story_points
                
                if direct_story_points > 0:
                    print(f"📋 Found {direct_story_points} story points from {len(direct_issues)} direct issues")
                    total_story_points += direct_story_points
            except Exception as e:
                print(f"⚠️  Could not load direct issues: {e}")
            
            if total_story_points > 0:
                print(f"✅ Total story points from JIRA: {total_story_points}")
            else:
                print("⚠️  No story points found in JIRA")
            
            return total_story_points
            
        except Exception as e:
            print(f"❌ Error loading story points from JIRA: {e}")
            return 0

    def load_story_points_interactive(self) -> float:
        """Interactive method to load story points from various sources"""
        print("\n📊 Story Points Loading Options")
        print("=" * 40)
        print("1. Load from JIRA (specific version/release)")
        print("2. Load from JIRA (all issues)")
        print("3. Load from Excel spec sheet")
        print("4. Enter manually")
        
        while True:
            choice = input("\nSelect option (1-4): ").strip()
            
            if choice == "1":
                if not self.jira_available:
                    print("❌ JIRA not available")
                    continue
                
                versions = self.jira_client.get_available_versions()
                if not versions:
                    print("⚠️  No versions found, loading all issues...")
                    return self.load_story_points_from_jira()
                
                print("\n📋 Available Versions:")
                for i, version in enumerate(versions, 1):
                    print(f"{i}. {version}")
                
                try:
                    version_choice = int(input(f"Select version (1-{len(versions)}): ").strip())
                    if 1 <= version_choice <= len(versions):
                        selected_version = versions[version_choice - 1]
                        return self.load_story_points_from_jira(selected_version)
                    else:
                        print("❌ Invalid choice")
                        continue
                except ValueError:
                    print("❌ Please enter a valid number")
                    continue
            
            elif choice == "2":
                return self.load_story_points_from_jira()  # No version filter
            
            elif choice == "3":
                return self.load_story_points_from_spec_sheet()
            
            elif choice == "4":
                try:
                    story_points = float(input("Enter total story points: ").strip())
                    return max(0, story_points)
                except ValueError:
                    print("❌ Please enter a valid number")
                    continue
            
            else:
                print("❌ Invalid option")

    def load_team_from_jira(self) -> bool:
        """Load team members from JIRA project (if available)"""
        if not self.jira_available:
            return False
        
        try:
            # This is a placeholder - in reality you'd query JIRA for project team members
            print("📊 Loading team information from JIRA...")
            # Could implement actual JIRA team loading here
            return False  # Not implemented yet
        except Exception as e:
            print(f"⚠️  Could not load team from JIRA: {e}")
            return False

    def get_team_interactive(self):
        """Interactive team member setup"""
        print("\n👥 Team Member Setup")
        print("=" * 40)
        
        while True:
            print("\nOptions:")
            print("1. Add team member")
            print("2. Use example team")
            print("3. Continue with current team")
            print("4. Show current team")
            
            choice = input("\nSelect option (1-4): ").strip()
            
            if choice == "1":
                self._add_team_member_interactive()
            elif choice == "2":
                self._add_example_team()
                break
            elif choice == "3":
                if self.team_members:
                    break
                else:
                    print("⚠️  No team members added yet!")
            elif choice == "4":
                self._show_current_team()
            else:
                print("❌ Invalid option")

    def _add_team_member_interactive(self):
        """Add a team member interactively"""
        print("\n📝 Add New Team Member")
        
        name = input("Name: ").strip()
        if not name:
            print("❌ Name required!")
            return
        
        role = input("Role: ").strip() or "Developer"
        
        try:
            fte = float(input("FTE (0.1-1.0): ").strip() or "1.0")
            fte = max(0.1, min(1.0, fte))  # Clamp between 0.1 and 1.0
        except ValueError:
            fte = 1.0
        
        try:
            sp_per_sprint = float(input("Story Points per Sprint: ").strip() or "8")
        except ValueError:
            sp_per_sprint = 8
        
        try:
            hourly_rate = float(input("Hourly Rate: ").strip() or "85")
        except ValueError:
            hourly_rate = 85
        
        seniority = input("Seniority (Junior/Mid/Senior/Lead): ").strip() or "Mid"
        skills_input = input("Skills (comma-separated): ").strip()
        skills = [skill.strip() for skill in skills_input.split(",")] if skills_input else []
        
        member = TeamMemberInfo(
            name=name,
            role=role,
            fte=fte,
            story_points_per_sprint=sp_per_sprint,
            hourly_rate=hourly_rate,
            seniority_level=seniority,
            skills=skills
        )
        
        self.add_team_member(member)

    def _add_example_team(self):
        """Add example team members"""
        print("📝 Adding example team...")
        
        self.add_team_member(TeamMemberInfo(
            name="Alice Johnson",
            role="Senior Developer",
            fte=1.0,
            story_points_per_sprint=10,
            hourly_rate=120,
            seniority_level="Senior",
            skills=["Python", "React", "AWS"]
        ))
        
        self.add_team_member(TeamMemberInfo(
            name="Bob Smith",
            role="Mid Developer",
            fte=1.0,
            story_points_per_sprint=7,
            hourly_rate=95,
            seniority_level="Mid",
            skills=["JavaScript", "Node.js", "Docker"]
        ))
        
        self.add_team_member(TeamMemberInfo(
            name="Carol Williams",
            role="Designer",
            fte=0.8,
            story_points_per_sprint=5,
            hourly_rate=85,
            seniority_level="Senior",
            skills=["UI/UX", "Figma", "Design Systems"]
        ))

    def _show_current_team(self):
        """Show current team members"""
        if not self.team_members:
            print("👥 No team members added yet")
            return
        
        print(f"\n👥 Current Team ({len(self.team_members)} members):")
        print("-" * 60)
        for i, member in enumerate(self.team_members, 1):
            print(f"{i}. {member.name} - {member.role}")
            print(f"   FTE: {member.fte}, SP/Sprint: {member.story_points_per_sprint}, Rate: €{member.hourly_rate}/h")
            print(f"   Seniority: {member.seniority_level}, Skills: {', '.join(member.skills) or 'None'}")

    def get_project_dates_interactive(self) -> Tuple[date, int]:
        """Interactive project date setup"""
        print("\n📅 Project Timeline Setup")
        print("=" * 40)
        
        # Get start date
        while True:
            date_input = input("Start date (YYYY-MM-DD) or 'next week': ").strip().lower()
            
            if date_input == "next week":
                start_date = date.today() + timedelta(days=7)
                break
            elif date_input == "today":
                start_date = date.today()
                break
            elif date_input == "tomorrow":
                start_date = date.today() + timedelta(days=1)
                break
            else:
                try:
                    start_date = datetime.strptime(date_input, "%Y-%m-%d").date()
                    break
                except ValueError:
                    print("❌ Invalid date format. Use YYYY-MM-DD, 'today', 'tomorrow', or 'next week'")
        
        # Get sprint length
        while True:
            sprint_input = input("Sprint length in weeks (default: 2): ").strip()
            
            if not sprint_input:
                sprint_weeks = 2
                break
            
            try:
                sprint_weeks = int(sprint_input)
                if 1 <= sprint_weeks <= 6:
                    break
                else:
                    print("❌ Sprint length must be between 1 and 6 weeks")
            except ValueError:
                print("❌ Please enter a valid number")
        
        print(f"📅 Project starts: {start_date}")
        print(f"⏱️  Sprint length: {sprint_weeks} weeks")
        
        return start_date, sprint_weeks

def main():
    """Main entry point for timeline generator"""
    print("📅 Project Timeline & Calendar Generator")
    print("=" * 60)
    print("🚀 Generate realistic project timelines with team availability and holidays")
    
    timeline = TimelineGenerator()
    
    try:
        # Load Dutch holidays for this year and next
        current_year = datetime.now().year
        timeline.load_dutch_holidays(current_year, current_year + 1)
        
        # Set up team interactively
        timeline.get_team_interactive()
        
        # Set up project dates
        start_date, sprint_weeks = timeline.get_project_dates_interactive()
        
        # Try to load story points from existing spec sheet
        print("\n📊 Loading Project Story Points")
        print("=" * 40)
        
        spec_sheet_options = ["spec-sheet.xlsx", "jira_cost_estimation.xlsx"]
        total_story_points = 0
        
        for spec_sheet in spec_sheet_options:
            if os.path.exists(spec_sheet):
                print(f"📋 Found spec sheet: {spec_sheet}")
                total_story_points = timeline.load_story_points_from_spec_sheet(spec_sheet)
                if total_story_points > 0:
                    break
        
        if total_story_points == 0:
            print("\n⚠️  No story points found in any spec sheet!")
            print("💡 This could be because:")
            print("   • The spec sheet doesn't contain story point values")
            print("   • The story point column header doesn't contain keywords: 'story point', 'sp', 'points'")
            print("   • Story points are in a different sheet or file")
            print("   • The JIRA sync hasn't been run to populate the spec sheet")
            print("\n📝 Please check the debug output above to see what was found.")
            
            while True:
                try:
                    manual_sp = input("\nEnter total story points manually: ").strip()
                    total_story_points = float(manual_sp)
                    if total_story_points > 0:
                        break
                    else:
                        print("❌ Story points must be greater than 0")
                except ValueError:
                    print("❌ Please enter a valid number")
        
        # Add custom holidays option
        print("\n🏖️  Custom Holidays")
        print("=" * 40)
        
        add_custom = input("Add custom holidays? (y/n): ").strip().lower()
        if add_custom in ['y', 'yes']:
            while True:
                holiday_date_input = input("Holiday date (YYYY-MM-DD) or 'done': ").strip()
                
                if holiday_date_input.lower() == 'done':
                    break
                
                try:
                    holiday_date = datetime.strptime(holiday_date_input, "%Y-%m-%d").date()
                    holiday_name = input("Holiday name: ").strip() or "Custom Holiday"
                    
                    affected_input = input("Affected team members (comma-separated, or 'all'): ").strip()
                    if affected_input.lower() in ['all', '']:
                        affected_members = []
                    else:
                        affected_members = [name.strip() for name in affected_input.split(",")]
                    
                    timeline.add_custom_holiday(HolidayInfo(
                        date=holiday_date,
                        name=holiday_name,
                        is_national=False,
                        affected_members=affected_members
                    ))
                    
                except ValueError:
                    print("❌ Invalid date format. Use YYYY-MM-DD")
        
        # Generate timeline
        print(f"\n⚡ Generating Timeline...")
        print("=" * 40)
        
        timeline.generate_sprint_timeline(start_date, total_story_points, sprint_weeks)
        
        # Save options
        print(f"\n💾 Save Timeline")
        print("=" * 40)
        
        custom_filename = input("Custom filename (or press Enter for auto): ").strip()
        filename = timeline.save_timeline_workbook(custom_filename if custom_filename else None)
        
        # Summary
        print(f"\n🎉 Timeline Generation Complete!")
        print("=" * 60)
        print(f"📊 Total Story Points: {total_story_points}")
        print(f"👥 Team Members: {len(timeline.team_members)}")
        print(f"📅 Sprints Planned: {len(timeline.sprints)}")
        print(f"🏖️  Holidays Tracked: {len(timeline.holidays)}")
        print(f"📅 Project Duration: {timeline.project_start_date} to {timeline.project_end_date}")
        print(f"💾 Saved To: {filename}")
        print(f"\n📋 Timeline includes:")
        print(f"   • Team member information and capacity")
        print(f"   • Dutch national holidays for {current_year}-{current_year + 1}")
        print(f"   • Sprint planning with realistic capacity calculations")
        print(f"   • Calendar view with holidays and sprint dates")
        print(f"   • Team availability considering FTE and working days")
        
        if timeline.sprints:
            avg_velocity = sum(s.team_velocity for s in timeline.sprints) / len(timeline.sprints)
            print(f"\n📈 Team Performance Metrics:")
            print(f"   • Average team velocity: {avg_velocity:.1f} SP/sprint")
            print(f"   • Estimated completion: {timeline.project_end_date}")
            print(f"   • Project duration: {(timeline.project_end_date - timeline.project_start_date).days} days")
        
    except KeyboardInterrupt:
        print("\n⏹️  Operation cancelled.")
    except Exception as e:
        print(f"\n💥 Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main() 