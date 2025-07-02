#!/usr/bin/env python3
"""
Spec Sheet Loading for Timeline Generation
Handles reading story points and project data from existing spec sheets
"""

import os
import openpyxl
import re


class SpecSheetLoader:
    """Loads project data from existing spec sheets"""
    
    @staticmethod
    def load_story_points_from_spec_sheet(spec_sheet_path: str = "output/spec-sheet.xlsx") -> float:
        """Load total story points from existing spec sheet (same as original)"""
        try:
            if not os.path.exists(spec_sheet_path):
                print(f"⚠️  Spec sheet not found: {spec_sheet_path}")
                return 0
            
            wb = openpyxl.load_workbook(spec_sheet_path, data_only=True)
            print(f"📊 Available sheets in {spec_sheet_path}: {wb.sheetnames}")
            
            # Try to find story points in various common sheet names
            sheet_names = ['Scope (Quantity)', 'Scope', 'Timeline', 'Summary', 'Sheet1']
            total_story_points = 0
            found_sheet = None
            
            for sheet_name in sheet_names:
                if sheet_name in wb.sheetnames:
                    found_sheet = sheet_name
                    break
            
            # If no common sheet names found, try the first sheet
            if not found_sheet and wb.sheetnames:
                found_sheet = wb.sheetnames[0]
                print(f"📊 Using first available sheet: {found_sheet}")
            
            if found_sheet:
                ws = wb[found_sheet]
                print(f"📊 Analyzing sheet: {found_sheet}")
                print(f"📊 Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
                
                # Look for story points columns
                story_point_cols = []
                header_info = {}
                
                print("\n🔍 Searching for story point columns...")
                for row_num in range(1, min(10, ws.max_row + 1)):  # Check first 10 rows for headers
                    for col_num in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.value and isinstance(cell.value, str):
                            cell_value_lower = cell.value.lower()
                            # Broader keyword search
                            keywords = ['story point', 'sp', 'points', 'story pts', 'storypoint', 'story_point']
                            for keyword in keywords:
                                if keyword in cell_value_lower:
                                    story_point_cols.append(col_num)
                                    header_info[col_num] = f"Row {row_num}: '{cell.value}'"
                                    print(f"   ✅ Found potential story point column: {col_num} (Row {row_num}: '{cell.value}')")
                                    break
                
                # Sum up story points from found columns
                for col in story_point_cols:
                    col_total = 0
                    print(f"\n📊 Processing column {col} ({header_info.get(col, 'Unknown header')}):")
                    
                    for row_num in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row_num, column=col).value
                        
                        # Handle different story point formats
                        story_point_value = None
                        
                        # Check if it's already a number
                        if isinstance(cell_value, (int, float)) and cell_value > 0:
                            story_point_value = cell_value
                        
                        # Check if it's text containing "Story Points: X"
                        elif isinstance(cell_value, str):
                            # Look for patterns like "Story Points: 5" or "SP: 3" etc.
                            patterns = [
                                r'story\s*points?\s*[:\-]?\s*(\d+(?:\.\d+)?)',
                                r'sp\s*[:\-]?\s*(\d+(?:\.\d+)?)',
                                r'points?\s*[:\-]?\s*(\d+(?:\.\d+)?)'
                            ]
                            
                            for pattern in patterns:
                                match = re.search(pattern, cell_value.lower())
                                if match:
                                    try:
                                        story_point_value = float(match.group(1))
                                        break
                                    except ValueError:
                                        continue
                        
                        if story_point_value is not None and story_point_value > 0:
                            col_total += story_point_value
                            total_story_points += story_point_value
                            print(f"   Row {row_num}: {story_point_value} SP (from '{cell_value}')")
                    
                    print(f"   Column {col} total: {col_total} SP")
                
                print(f"\n📋 Total story points found: {total_story_points}")
            
            if total_story_points > 0:
                print(f"✅ Successfully loaded {total_story_points} story points from {spec_sheet_path}")
                return total_story_points
            else:
                print("⚠️  No story points found in spec sheet")
                return 0
            
        except Exception as e:
            print(f"❌ Error reading spec sheet: {e}")
            return 0 