#!/usr/bin/env python3
"""
Excel Workbook Generation for Timeline
Handles creating and formatting Excel workbooks with timeline data
"""

from datetime import datetime
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelGenerator:
    """Generates Excel workbooks with timeline and team data"""
    
    def __init__(self, team_members: List, sprints: List, 
                 holiday_manager, calendar_generator,
                 project_start_date: Optional = None, 
                 project_end_date: Optional = None):
        self.team_members = team_members
        self.sprints = sprints
        self.holiday_manager = holiday_manager
        self.calendar_generator = calendar_generator
        self.project_start_date = project_start_date
        self.project_end_date = project_end_date
    
    def create_timeline_workbook(self, filename: str = None) -> str:
        """Create the complete timeline workbook with all sheets"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"project_timeline_{timestamp}.xlsx"
        
        print(f"📊 Creating timeline workbook: {filename}")
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create all sheets
        self._create_team_members_sheet(wb)
        self._create_holidays_sheet(wb)
        self._create_sprint_timeline_sheet(wb)
        self.calendar_generator.create_calendar_view_sheet(wb, self.project_start_date, self.project_end_date)
        
        # Save workbook
        wb.save(filename)
        print(f"✅ Timeline workbook saved: {filename}")
        
        return filename
    
    def _create_team_members_sheet(self, wb: openpyxl.Workbook):
        """Create team members sheet with enhanced information"""
        ws = wb.create_sheet("Team Members")
        
        # Headers
        headers = ["Name", "Role", "Availability", "SP/Sprint", "Hourly Rate", "Email", "Holidays"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Team member data
        for row, member in enumerate(self.team_members, 2):
            ws.cell(row=row, column=1, value=member.name)
            ws.cell(row=row, column=2, value=member.role)
            ws.cell(row=row, column=3, value=f"{member.availability*100:.0f}%")
            ws.cell(row=row, column=4, value=member.story_points_per_sprint)
            ws.cell(row=row, column=5, value=f"€{member.hourly_rate:.2f}")
            ws.cell(row=row, column=6, value=member.email or "")
            ws.cell(row=row, column=7, value=len(member.holidays))
        
        # Auto-size columns
        self._auto_size_columns(ws)
    
    def _create_holidays_sheet(self, wb: openpyxl.Workbook):
        """Create holidays sheet"""
        ws = wb.create_sheet("Holidays")
        
        # Headers
        headers = ["Date", "Name", "Type", "Affected Members"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Holiday data
        sorted_holidays = sorted(self.holiday_manager.holidays, key=lambda h: h.date)
        for row, holiday in enumerate(sorted_holidays, 2):
            ws.cell(row=row, column=1, value=holiday.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=holiday.name)
            ws.cell(row=row, column=3, value="National" if holiday.is_national else "Personal")
            ws.cell(row=row, column=4, value=", ".join(holiday.affected_members) if holiday.affected_members else "Everyone")
        
        # Auto-size columns
        self._auto_size_columns(ws)
    
    def _create_sprint_timeline_sheet(self, wb: openpyxl.Workbook):
        """Create sprint timeline sheet (main timeline view)"""
        ws = wb.create_sheet("Sprint Timeline")
        
        # Headers
        headers = ["Sprint", "Start Date", "End Date", "Story Points", "Team Velocity", "Duration (Days)", "Capacity Utilization"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Sprint data
        for row, sprint in enumerate(self.sprints, 2):
            duration = (sprint.end_date - sprint.start_date).days + 1
            utilization = (sprint.story_points / sprint.team_velocity * 100) if sprint.team_velocity > 0 else 0
            
            ws.cell(row=row, column=1, value=f"Sprint {sprint.number}")
            ws.cell(row=row, column=2, value=sprint.start_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=3, value=sprint.end_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=4, value=sprint.story_points)
            ws.cell(row=row, column=5, value=sprint.team_velocity)
            ws.cell(row=row, column=6, value=duration)
            ws.cell(row=row, column=7, value=f"{utilization:.1f}%")
        
        # Summary row
        if self.sprints:
            summary_row = len(self.sprints) + 3
            ws.cell(row=summary_row, column=1, value="TOTAL")
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            ws.cell(row=summary_row, column=4, value=sum(s.story_points for s in self.sprints))
            if self.project_start_date and self.project_end_date:
                ws.cell(row=summary_row, column=6, value=(self.project_end_date - self.project_start_date).days + 1)
        
        # Auto-size columns
        self._auto_size_columns(ws)
    
    def _auto_size_columns(self, ws):
        """Auto-size columns for better readability"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width 