#!/usr/bin/env python3
"""
Calendar View Generation for Timeline
Handles creating monthly calendar views with project timeline data
"""

import calendar
from datetime import date, timedelta
from typing import List, Optional
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font


class CalendarGenerator:
    """Generates calendar views for timeline visualization"""
    
    def __init__(self, holiday_manager, sprints: List):
        self.holiday_manager = holiday_manager
        self.sprints = sprints
    
    def create_calendar_view_sheet(self, workbook: openpyxl.Workbook, 
                                 project_start_date: Optional[date], 
                                 project_end_date: Optional[date]):
        """Create a calendar view showing team availability and project timeline"""
        ws = workbook.create_sheet("Calendar View")
        
        if not project_start_date or not project_end_date:
            ws.cell(row=1, column=1, value="No project timeline generated yet")
            return
        
        # Title
        ws.cell(row=1, column=1, value=f"Project Calendar: {project_start_date} to {project_end_date}")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        
        # Legend
        ws.cell(row=2, column=1, value="Legend:")
        ws.cell(row=2, column=2, value="Sprint Days")
        ws.cell(row=2, column=2).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        ws.cell(row=2, column=3, value="Holidays")
        ws.cell(row=2, column=3).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        ws.cell(row=2, column=4, value="Weekends")
        ws.cell(row=2, column=4).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Generate monthly calendar views
        current_month = project_start_date.replace(day=1)
        end_month = project_end_date.replace(day=1)
        row_offset = 5
        
        while current_month <= end_month:
            row_offset = self._create_monthly_calendar(ws, current_month, row_offset)
            # Move to next month
            if current_month.month == 12:
                current_month = current_month.replace(year=current_month.year + 1, month=1)
            else:
                current_month = current_month.replace(month=current_month.month + 1)
    
    def _create_monthly_calendar(self, ws: openpyxl.Workbook, month_date: date, start_row: int) -> int:
        """Create a single month calendar view"""
        month_name = month_date.strftime("%B %Y")
        
        # Month header
        ws.cell(row=start_row, column=1, value=month_name)
        ws.cell(row=start_row, column=1).font = Font(size=12, bold=True)
        
        # Day headers
        day_headers = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for col, day in enumerate(day_headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=day)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Get calendar for this month
        cal = calendar.monthcalendar(month_date.year, month_date.month)
        
        # Fill in the calendar
        for week_num, week in enumerate(cal):
            row = start_row + 2 + week_num
            for day_num, day in enumerate(week):
                col = day_num + 1
                
                if day == 0:
                    continue  # Empty cell for days not in this month
                
                current_date = date(month_date.year, month_date.month, day)
                cell = ws.cell(row=row, column=col, value=day)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Check if this date is special
                cell_fill = None
                
                # Check for holidays
                if self.holiday_manager.is_holiday(current_date):
                    cell_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                # Check for sprint dates
                for sprint in self.sprints:
                    if sprint.start_date <= current_date <= sprint.end_date:
                        if cell_fill is None:  # Don't override holiday colors
                            cell_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        break
                
                # Check for weekends
                if current_date.weekday() >= 5:  # Saturday = 5, Sunday = 6
                    if cell_fill is None:
                        cell_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                
                if cell_fill:
                    cell.fill = cell_fill
                
                # Add border
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Set column widths
        for col in range(1, 8):
            column_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[column_letter].width = 6
        
        return start_row + len(cal) + 4  # Return next available row 