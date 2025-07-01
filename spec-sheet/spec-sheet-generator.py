#!/usr/bin/env python3
"""
Enhanced Jira to Spec Sheet Sync Tool
Syncs Jira epics and stories with the existing sophisticated pricing structure
Now includes sprint planning and team composition features
"""

import sys
import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from utils.config import JiraConfig, SpreadsheetConfig
from utils.jira_client import JiraClient
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import math
import json

# Team composition and sprint planning classes
class TeamMember:
    """Represents a team member with their role and capacity"""
    def __init__(self, role: str, fte: float, story_points_per_sprint: float, hourly_rate: float = 95.37):
        self.role = role
        self.fte = fte  # Full Time Equivalent (0.0 to 1.0)
        self.base_story_points_per_sprint = story_points_per_sprint
        self.hourly_rate = hourly_rate
        self.effective_velocity = story_points_per_sprint * fte
    
    def __str__(self):
        return f"{self.role} ({self.fte} FTE): {self.effective_velocity:.1f} SP/sprint"

class Team:
    """Represents a development team with various roles and capacities"""
    def __init__(self, name: str = "Development Team", settings_config: Dict = None):
        self.name = name
        self.members: List[TeamMember] = []
        
        # Use configurable settings or defaults
        if settings_config:
            self.sprint_length_weeks = settings_config["sprint_planning"]["default_sprint_length_weeks"]
            self.sprint_overhead = settings_config["sprint_planning"]["sprint_overhead_percentage"]
        else:
            self.sprint_length_weeks = 2  # Default 2-week sprints
            self.sprint_overhead = 0.15   # 15% overhead for meetings, planning, etc.
    
    def add_member(self, member: TeamMember):
        """Add a team member"""
        self.members.append(member)
    
    def get_total_velocity(self) -> float:
        """Calculate total team velocity per sprint"""
        total_velocity = sum(member.effective_velocity for member in self.members)
        # Apply sprint overhead reduction
        return total_velocity * (1 - self.sprint_overhead)
    
    def get_total_fte(self) -> float:
        """Get total FTE of the team"""
        return sum(member.fte for member in self.members)
    
    def get_composition_summary(self) -> str:
        """Get a summary of team composition"""
        if not self.members:
            return "Empty team"
        
        summary = f"{self.name} ({self.get_total_fte():.1f} total FTE):\n"
        for member in self.members:
            summary += f"  • {member}\n"
        summary += f"Total velocity: {self.get_total_velocity():.1f} SP/sprint"
        return summary

class EnhancedSpecSheetSync:
    """Enhanced class for synchronizing Jira with the sophisticated spec sheet structure"""
    
    def __init__(self):
        try:
            self.jira_config = JiraConfig()
            self.jira_client = JiraClient(self.jira_config)
            
            # Load existing spec sheet structure
            self.spec_sheet_path = "spec-sheet.xlsx"
            self.dod_config_path = "dod_config.json"
            self.settings_config_path = "settings_config.json"
            
            # Load configurations from files
            self.settings_config = self._load_settings_config()
            self.dod_config = self._load_dod_config()
            
            # Extract commonly used settings for easy access
            self.base_story_point_price = self.settings_config["pricing"]["base_story_point_price"]
            self.experimental_variance = self.settings_config["pricing"]["experimental_variance"]
            # Calculate hourly rate dynamically
            pricing_config = self.settings_config["pricing"]
            self.hourly_rate = pricing_config["base_hourly_rate"] * (1 - pricing_config["hourly_rate_discount"])
            
            # DoD impact will be calculated after loading the structure
            
            # Type of work field configuration
            self.type_of_work_field = self.jira_config.type_of_work_field
            self.risk_priority = self.settings_config["risk_assessment"]["risk_priority"]
            
            # Sprint planning configuration
            self.sprint_planning_enabled = True
            self.default_teams = self._create_default_teams()
            
        except Exception as e:
            print(f"❌ Configuration error: {e}")
            sys.exit(1)
    
    def _create_default_teams(self) -> Dict[str, Team]:
        """Create default team compositions for analysis using configurable settings"""
        teams = {}
        team_defaults = self.settings_config["team_defaults"]
        
        # Small team
        small_team = Team("Small Team", self.settings_config)
        small_team.add_member(TeamMember("Senior Developer", 1.0, team_defaults["senior_developer_velocity"]))
        small_team.add_member(TeamMember("Junior Developer", 1.0, team_defaults["junior_developer_velocity"]))
        small_team.add_member(TeamMember("Designer", 0.5, team_defaults["designer_velocity"]))
        teams["small"] = small_team
        
        # Medium team
        medium_team = Team("Medium Team", self.settings_config)
        medium_team.add_member(TeamMember("Tech Lead", 1.0, team_defaults["tech_lead_velocity"]))
        medium_team.add_member(TeamMember("Senior Developer", 2.0, team_defaults["senior_developer_velocity"]))
        medium_team.add_member(TeamMember("Junior Developer", 1.0, team_defaults["junior_developer_velocity"]))
        medium_team.add_member(TeamMember("Designer", 1.0, team_defaults["designer_velocity"]))
        medium_team.add_member(TeamMember("QA Engineer", 0.5, team_defaults["qa_engineer_velocity"]))
        teams["medium"] = medium_team
        
        # Large team
        large_team = Team("Large Team", self.settings_config)
        large_team.add_member(TeamMember("Tech Lead", 1.0, team_defaults["tech_lead_velocity"]))
        large_team.add_member(TeamMember("Senior Developer", 3.0, team_defaults["senior_developer_velocity"]))
        large_team.add_member(TeamMember("Mid Developer", 2.0, team_defaults["mid_developer_velocity"]))
        large_team.add_member(TeamMember("Junior Developer", 2.0, team_defaults["junior_developer_velocity"]))
        large_team.add_member(TeamMember("Designer", 1.0, team_defaults["designer_velocity"]))
        large_team.add_member(TeamMember("QA Engineer", 1.0, team_defaults["qa_engineer_velocity"]))
        large_team.add_member(TeamMember("DevOps Engineer", 0.5, team_defaults["devops_engineer_velocity"]))
        teams["large"] = large_team
        
        return teams
    
    def _load_settings_config(self) -> Dict:
        """Load settings configuration from JSON file"""
        try:
            if os.path.exists(self.settings_config_path):
                with open(self.settings_config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                print(f"✅ Loaded settings configuration from {self.settings_config_path}")
                return config
            else:
                print(f"⚠️  Settings config file not found at {self.settings_config_path}, using defaults")
                return self._get_default_settings_config()
        except Exception as e:
            print(f"❌ Error loading settings config: {e}, using defaults")
            return self._get_default_settings_config()
    
    def _get_default_settings_config(self) -> Dict:
        """Get default settings configuration if file doesn't exist"""
        return {
            "description": "Default settings configuration",
            "version": "1.0",
            "pricing": {
                "base_story_point_price": 130,
                "experimental_variance": 0.3,
                "base_hourly_rate": 127.16,
                "hourly_rate_discount": 0.25
            },
            "sprint_planning": {
                "default_sprint_length_weeks": 2,
                "sprint_overhead_percentage": 0.15,
                "working_days_per_week": 5,
                "hours_per_story_point": 8
            },
            "team_defaults": {
                "default_hourly_rate": 95.37,
                "senior_developer_velocity": 8,
                "mid_developer_velocity": 6,
                "junior_developer_velocity": 5,
                "tech_lead_velocity": 10,
                "designer_velocity": 6,
                "qa_engineer_velocity": 4,
                "devops_engineer_velocity": 7
            },
            "risk_assessment": {
                "proven_threshold_story_points": 3,
                "experimental_threshold_story_points": 8,
                "risk_priority": {
                    "proven": 1,
                    "experimental": 2,
                    "dependant": 3,
                    "dependent": 3
                }
            },
            "recommendations": {
                "large_project_threshold": 100,
                "small_project_threshold": 20,
                "significant_sprints_saved_threshold": 4,
                "high_experimental_percentage": 30,
                "significant_dependent_percentage": 20
            },
            "ui_formatting": {
                "header_color": "366092",
                "summary_header_color": "366092",
                "epic_background_color": "E6F3FF",
                "section_background_color": "D4EDDA",
                "default_column_widths": [50, 20, 20, 30, 15, 12, 15, 12, 12, 15, 20],
                "header_row_height": 80
            }
        }
    
    def _load_dod_config(self) -> Dict:
        """Load Definition of Done configuration from JSON file"""
        try:
            if os.path.exists(self.dod_config_path):
                with open(self.dod_config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                print(f"✅ Loaded DoD configuration from {self.dod_config_path}")
                return config
            else:
                print(f"⚠️  DoD config file not found at {self.dod_config_path}, using defaults")
                return self._get_default_dod_config()
        except Exception as e:
            print(f"❌ Error loading DoD config: {e}, using defaults")
            return self._get_default_dod_config()
    
    def _get_default_dod_config(self) -> Dict:
        """Get default DoD configuration if file doesn't exist"""
        return {
            "description": "Default Definition of Done configuration",
            "version": "1.0",
            "categories": [
                {
                    "name": "Code quality & documentation",
                    "items": [
                        {
                            "description": "Code is structured, modular, and follows best practices",
                            "moscow": "Must Have",
                            "impact_points": "8",
                            "impact_percentage": 0.04
                        },
                        {
                            "description": "Code is reviewed and approved via pull requests",
                            "moscow": "Must Have",
                            "impact_points": "5",
                            "impact_percentage": 0.03
                        }
                    ]
                },
                {
                    "name": "Performance & optimization",
                    "items": [
                        {
                            "description": "Performance is optimized for production use",
                            "moscow": "Should Have",
                            "impact_points": "13",
                            "impact_percentage": 0.07
                        }
                    ]
                }
            ]
        }
    
    def calculate_sprint_estimates(self, total_story_points: float, team: Team) -> Dict[str, any]:
        """Calculate sprint estimates for a given team and story points"""
        team_velocity = team.get_total_velocity()
        
        if team_velocity <= 0:
            return {
                'sprints': 0,
                'weeks': 0,
                'months': 0,
                'team_velocity': 0,
                'error': 'Team has no effective velocity'
            }
        
        sprints_needed = math.ceil(total_story_points / team_velocity)
        weeks_needed = sprints_needed * team.sprint_length_weeks
        months_needed = weeks_needed / 4.33  # Average weeks per month
        
        return {
            'sprints': sprints_needed,
            'weeks': weeks_needed,
            'months': months_needed,
            'team_velocity': team_velocity,
            'team_composition': team.get_composition_summary(),
            'total_story_points': total_story_points,
            'story_points_per_sprint': team_velocity
        }
    
    def compare_team_options(self, total_story_points: float) -> Dict[str, Dict[str, any]]:
        """Compare different team compositions and their impact on timeline"""
        comparison = {}
        
        print(f"\n📊 Sprint Planning Analysis for {total_story_points} story points")
        print("=" * 60)
        
        for team_name, team in self.default_teams.items():
            estimates = self.calculate_sprint_estimates(total_story_points, team)
            comparison[team_name] = estimates
            
            print(f"\n{team.name}:")
            print(f"  Team Size: {team.get_total_fte():.1f} FTE")
            print(f"  Velocity: {estimates['team_velocity']:.1f} SP/sprint")
            print(f"  Duration: {estimates['sprints']} sprints ({estimates['weeks']} weeks, {estimates['months']:.1f} months)")
            print(f"  Team: {len(team.members)} people")
        
        return comparison
    
    def create_custom_team(self, team_composition: List[Dict]) -> Team:
        """Create a custom team from a list of team member specifications"""
        custom_team = Team("Custom Team")
        
        for member_spec in team_composition:
            role = member_spec.get('role', 'Developer')
            fte = member_spec.get('fte', 1.0)
            velocity = member_spec.get('story_points_per_sprint', 6)
            hourly_rate = member_spec.get('hourly_rate', 95.37)
            
            member = TeamMember(role, fte, velocity, hourly_rate)
            custom_team.add_member(member)
        
        return custom_team
    
    def generate_sprint_planning_report(self, epics: List[Dict] = None) -> Dict[str, any]:
        """Generate comprehensive sprint planning report"""
        if epics is None:
            # Use pre-selected epics if available, otherwise fetch all
            if hasattr(self, 'selected_epics') and self.selected_epics:
                epics = self.selected_epics
            else:
                epics = self.jira_client.get_epics()
        
        # Calculate total story points across all epics
        total_story_points = 0
        epic_breakdown = {}
        
        for epic in epics:
            epic_key = epic['key']
            stories = self.jira_client.get_stories_for_epic(epic_key)
            epic_points = sum(self.jira_client.get_story_points(story) or 0 for story in stories)
            epic_breakdown[epic_key] = {
                'summary': epic['fields']['summary'],
                'story_points': epic_points,
                'story_count': len(stories)
            }
            total_story_points += epic_points
        
        # Compare team options
        team_comparison = self.compare_team_options(total_story_points)
        
        # Risk analysis based on story types
        risk_breakdown = self._analyze_risk_distribution(epics)
        
        return {
            'total_story_points': total_story_points,
            'epic_breakdown': epic_breakdown,
            'team_comparison': team_comparison,
            'risk_breakdown': risk_breakdown,
            'recommendations': self._generate_recommendations(total_story_points, team_comparison, risk_breakdown)
        }
    
    def _analyze_risk_distribution(self, epics: List[Dict]) -> Dict[str, any]:
        """Analyze the distribution of risk profiles across stories"""
        risk_counts = {'proven': 0, 'experimental': 0, 'dependant': 0}
        risk_story_points = {'proven': 0, 'experimental': 0, 'dependant': 0}
        
        for epic in epics:
            stories = self.jira_client.get_stories_for_epic(epic['key'])
            for story in stories:
                risk_profile = self.determine_risk_profile(story)
                story_points = self.jira_client.get_story_points(story) or 0
                
                risk_counts[risk_profile] += 1
                risk_story_points[risk_profile] += story_points
        
        total_stories = sum(risk_counts.values())
        total_points = sum(risk_story_points.values())
        
        return {
            'counts': risk_counts,
            'story_points': risk_story_points,
            'percentages': {
                risk: (count / total_stories * 100) if total_stories > 0 else 0
                for risk, count in risk_counts.items()
            },
            'story_point_percentages': {
                risk: (points / total_points * 100) if total_points > 0 else 0
                for risk, points in risk_story_points.items()
            }
        }
    
    def _generate_recommendations(self, total_story_points: float, team_comparison: Dict, risk_breakdown: Dict) -> List[str]:
        """Generate recommendations based on analysis using configurable thresholds"""
        recommendations = []
        rec_config = self.settings_config["recommendations"]
        sprint_config = self.settings_config["sprint_planning"]
        
        # Timeline recommendations
        small_sprints = team_comparison.get('small', {}).get('sprints', 0)
        large_sprints = team_comparison.get('large', {}).get('sprints', 0)
        
        if small_sprints > 0 and large_sprints > 0:
            time_saved = small_sprints - large_sprints
            if time_saved > rec_config["significant_sprints_saved_threshold"]:
                weeks_saved = time_saved * sprint_config["default_sprint_length_weeks"]
                recommendations.append(f"Consider larger team - saves {time_saved} sprints ({weeks_saved} weeks)")
        
        # Risk-based recommendations
        experimental_pct = risk_breakdown.get('story_point_percentages', {}).get('experimental', 0)
        dependant_pct = risk_breakdown.get('story_point_percentages', {}).get('dependant', 0)
        
        if experimental_pct > rec_config["high_experimental_percentage"]:
            recommendations.append(f"High experimental work (>{rec_config['high_experimental_percentage']}%) - consider prototyping phase")
        
        if dependant_pct > rec_config["significant_dependent_percentage"]:
            recommendations.append(f"Significant dependent work (>{rec_config['significant_dependent_percentage']}%) - ensure external dependencies are ready")
        
        # Story points recommendations
        if total_story_points > rec_config["large_project_threshold"]:
            recommendations.append("Large project - consider breaking into phases or releases")
        
        if total_story_points < rec_config["small_project_threshold"]:
            recommendations.append("Small project - single developer might be sufficient")
        
        return recommendations

    def add_sprint_planning_sheet(self):
        """Add a new sheet for sprint planning analysis"""
        sheet_name = 'Sprint Planning'
        
        # Remove existing sheet if it exists
        if sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet_name])
        
        # Create new sheet
        ws = self.workbook.create_sheet(sheet_name)
        
        # Generate sprint planning report
        report = self.generate_sprint_planning_report()
        
        # Set up headers and content
        self._setup_sprint_planning_sheet(ws, report)
        
        # Save workbook
        self.workbook.save(self.spec_sheet_path)
        print(f"✅ Added sprint planning analysis to {sheet_name} sheet")
    
    def _setup_sprint_planning_sheet(self, ws, report: Dict):
        """Set up the sprint planning sheet with analysis data"""
        current_row = 1
        
        # Title
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = "Sprint Planning & Team Composition Analysis"
        title_cell.font = Font(bold=True, size=16)
        current_row += 3
        
        # Project Summary
        ws.cell(row=current_row, column=1).value = "Project Overview"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Total Story Points:"
        ws.cell(row=current_row, column=2).value = report['total_story_points']
        current_row += 1
        
        ws.cell(row=current_row, column=1).value = "Number of Epics:"
        ws.cell(row=current_row, column=2).value = len(report['epic_breakdown'])
        current_row += 2
        
        # Epic Breakdown
        ws.cell(row=current_row, column=1).value = "Epic Breakdown"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1
        
        # Epic headers
        headers = ["Epic Key", "Summary", "Story Points", "Story Count"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=i)
            cell.value = header
            cell.font = Font(bold=True)
        current_row += 1
        
        # Epic data
        for epic_key, epic_data in report['epic_breakdown'].items():
            ws.cell(row=current_row, column=1).value = epic_key
            ws.cell(row=current_row, column=2).value = epic_data['summary']
            ws.cell(row=current_row, column=3).value = epic_data['story_points']
            ws.cell(row=current_row, column=4).value = epic_data['story_count']
            current_row += 1
        
        current_row += 2
        
        # Team Comparison
        ws.cell(row=current_row, column=1).value = "Team Composition Comparison"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1
        
        # Team comparison headers
        team_headers = ["Team Size", "Total FTE", "Velocity (SP/sprint)", "Sprints Needed", "Weeks", "Months"]
        for i, header in enumerate(team_headers, 2):
            cell = ws.cell(row=current_row, column=i)
            cell.value = header
            cell.font = Font(bold=True)
        current_row += 1
        
        # Team comparison data
        for team_name, estimates in report['team_comparison'].items():
            ws.cell(row=current_row, column=1).value = team_name.title()
            ws.cell(row=current_row, column=2).value = len(self.default_teams[team_name].members)
            ws.cell(row=current_row, column=3).value = f"{self.default_teams[team_name].get_total_fte():.1f}"
            ws.cell(row=current_row, column=4).value = f"{estimates['team_velocity']:.1f}"
            ws.cell(row=current_row, column=5).value = estimates['sprints']
            ws.cell(row=current_row, column=6).value = estimates['weeks']
            ws.cell(row=current_row, column=7).value = f"{estimates['months']:.1f}"
            current_row += 1
        
        current_row += 2
        
        # Risk Analysis
        ws.cell(row=current_row, column=1).value = "Risk Profile Analysis"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1
        
        # Risk headers
        risk_headers = ["Risk Profile", "Stories", "Story Points", "% of Stories", "% of Story Points"]
        for i, header in enumerate(risk_headers, 1):
            cell = ws.cell(row=current_row, column=i)
            cell.value = header
            cell.font = Font(bold=True)
        current_row += 1
        
        # Risk data
        for risk_profile in ['proven', 'experimental', 'dependant']:
            ws.cell(row=current_row, column=1).value = risk_profile.title()
            ws.cell(row=current_row, column=2).value = report['risk_breakdown']['counts'][risk_profile]
            ws.cell(row=current_row, column=3).value = report['risk_breakdown']['story_points'][risk_profile]
            ws.cell(row=current_row, column=4).value = f"{report['risk_breakdown']['percentages'][risk_profile]:.1f}%"
            ws.cell(row=current_row, column=5).value = f"{report['risk_breakdown']['story_point_percentages'][risk_profile]:.1f}%"
            current_row += 1
        
        current_row += 2
        
        # Recommendations
        if report['recommendations']:
            ws.cell(row=current_row, column=1).value = "Recommendations"
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
            current_row += 1
            
            for i, recommendation in enumerate(report['recommendations'], 1):
                ws.cell(row=current_row, column=1).value = f"{i}. {recommendation}"
                current_row += 1
        
        # Adjust column widths
        column_widths = [25, 50, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    def load_spec_sheet_structure(self):
        """Load the existing spec sheet or create a new one"""
        try:
            if os.path.exists(self.spec_sheet_path):
                self.workbook = openpyxl.load_workbook(self.spec_sheet_path)
                
                # Load DoD impacts
                self.dod_impacts = self._load_dod_impacts()
                
                # Calculate total DoD impact from loaded impacts
                self.dod_impact_total = self._calculate_dod_impact_total()
                
                # Load settings
                self.settings = self._load_settings()
                
                print("✅ Loaded existing spec sheet structure")
            else:
                print(f"📝 Spec sheet not found at {self.spec_sheet_path}, creating new one...")
                self._create_new_spec_sheet()
                print("✅ Created new spec sheet with default structure")
            
            return True
            
        except Exception as e:
            print(f"❌ Error with spec sheet: {e}")
            return False
    
    def _load_dod_impacts(self) -> Dict[str, float]:
        """Load Definition of Done impacts from the spec sheet"""
        dod_impacts = {}
        try:
            df = pd.read_excel(self.spec_sheet_path, sheet_name='Definition of Done (Quality)')
            
            for idx, row in df.iterrows():
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[3]):
                    description = str(row.iloc[0])
                    # Skip category headers and sum row
                    if description in ['Sum'] or description.endswith('&') or description.endswith('compatibility') or description.endswith('deployment'):
                        continue
                    
                    try:
                        impact = float(row.iloc[3])
                        # Create a clean key from the description
                        clean_key = description.lower().replace(' ', '_').replace('(', '').replace(')', '').replace(',', '').replace('-', '_')
                        dod_impacts[clean_key] = impact
                        print(f"   📋 Loaded DoD: {description} = {impact:.1%}")
                    except (ValueError, TypeError):
                        continue
                        
        except Exception as e:
            print(f"Warning: Could not load DoD impacts, using defaults: {e}")
            # Return default impacts if loading fails
            return self._get_default_dod_impacts()
            
        if not dod_impacts:
            print("Warning: No DoD impacts found in sheet, using defaults")
            return self._get_default_dod_impacts()
            
        return dod_impacts
    
    def _load_settings(self) -> Dict[str, any]:
        """Load settings from the spec sheet, with fallback to configuration file"""
        pricing_config = self.settings_config["pricing"]
        # Calculate hourly rate dynamically
        calculated_hourly_rate = pricing_config["base_hourly_rate"] * (1 - pricing_config["hourly_rate_discount"])
        settings = {
            'base_story_point_price': pricing_config["base_story_point_price"],
            'experimental_variance': pricing_config["experimental_variance"],
            'hourly_rate': calculated_hourly_rate
        }
        
        try:
            df = pd.read_excel(self.spec_sheet_path, sheet_name='Settings')
            
            for idx, row in df.iterrows():
                if pd.notna(row.iloc[0]):
                    setting_name = str(row.iloc[0]).lower()
                    if 'base price of 1 story point' in setting_name and pd.notna(row.iloc[1]):
                        settings['base_story_point_price'] = float(row.iloc[1])
                    elif 'experimental' in setting_name and 'range' in setting_name and pd.notna(row.iloc[1]):
                        settings['experimental_variance'] = float(row.iloc[1])
                        
        except Exception as e:
            print(f"Warning: Could not load settings: {e}")
            
        return settings
    
    def _create_new_spec_sheet(self):
        """Create a new spec sheet with all required sheets"""
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Create Scope (Quantity) sheet
        scope_sheet = self.workbook.create_sheet('Scope (Quantity)')
        
        # Create Definition of Done sheet with default values
        dod_sheet = self.workbook.create_sheet('Definition of Done (Quality)')
        self._setup_default_dod_sheet(dod_sheet)
        
        # Create Settings sheet with default values
        settings_sheet = self.workbook.create_sheet('Settings')
        self._setup_default_settings_sheet(settings_sheet)
        
        # Set default values
        self.dod_impacts = self._get_default_dod_impacts()
        
        # Calculate total DoD impact from default impacts
        self.dod_impact_total = self._calculate_dod_impact_total()
        
        self.settings = self._load_settings()
        
        # Save the new workbook
        self.workbook.save(self.spec_sheet_path)
    
    def _setup_default_dod_sheet(self, ws):
        """Set up Definition of Done sheet using configuration"""
        headers = [
            "Definition of Done",
            "MoSCoW",
            "Price Impact",
            "Price Impact %"
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i).value = header
            ws.cell(row=1, column=i).font = Font(bold=True)
        
        current_row = 2
        all_items = []
        
        # Use configuration to build DoD sheet
        for category in self.dod_config.get('categories', []):
            category_name = category.get('name', 'Unknown Category')
            
            # Add category header
            ws.cell(row=current_row, column=1).value = category_name
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            # Add items for this category
            for item in category.get('items', []):
                desc = item.get('description', '')
                moscow = item.get('moscow', 'Should Have')
                impact_desc = item.get('impact_points', '')
                impact_pct = item.get('impact_percentage', 0.0)
                
                ws.cell(row=current_row, column=1).value = desc
                ws.cell(row=current_row, column=2).value = moscow
                ws.cell(row=current_row, column=3).value = impact_desc
                ws.cell(row=current_row, column=4).value = impact_pct
                
                all_items.append(impact_pct)
                current_row += 1
            
            # Add empty row between categories
            current_row += 1
        
        # Add sum row with calculated total
        current_row += 1
        ws.cell(row=current_row, column=3).value = "Sum"
        ws.cell(row=current_row, column=3).font = Font(bold=True)
        
        # Calculate sum from all items
        calculated_sum = sum(all_items)
        ws.cell(row=current_row, column=4).value = calculated_sum
        ws.cell(row=current_row, column=4).font = Font(bold=True)
        
        print(f"📊 Generated DoD sheet with {len(all_items)} items, total impact: {calculated_sum:.1%}")
    
    def _setup_default_settings_sheet(self, ws):
        """Set up default Settings sheet using configurable values"""
        pricing_config = self.settings_config["pricing"]
        # Calculate hourly rate dynamically
        calculated_hourly_rate = pricing_config["base_hourly_rate"] * (1 - pricing_config["hourly_rate_discount"])
        settings_data = [
            ("Range for Experimental pricing (up & down)", pricing_config["experimental_variance"]),
            ("Base price of 8 story points", pricing_config["base_story_point_price"] * 8),
            ("Base price of 1 story point", pricing_config["base_story_point_price"]),
            ("Hourly rate", calculated_hourly_rate),
        ]
        
        ws.cell(row=1, column=1).value = "Settings"
        ws.cell(row=1, column=1).font = Font(bold=True)
        
        for i, (setting, value) in enumerate(settings_data, 2):
            ws.cell(row=i, column=1).value = setting
            ws.cell(row=i, column=2).value = value
    
    def _get_default_dod_impacts(self):
        """Get DoD impacts from loaded configuration"""
        dod_impacts = {}
        
        for category in self.dod_config.get('categories', []):
            for item in category.get('items', []):
                description = item.get('description', '')
                impact_percentage = item.get('impact_percentage', 0.0)
                
                # Create a clean key from the description
                clean_key = description.lower().replace(' ', '_').replace('(', '').replace(')', '').replace(',', '').replace('-', '_').replace('&', 'and')
                dod_impacts[clean_key] = impact_percentage
        
        print(f"📋 Loaded {len(dod_impacts)} DoD impacts from configuration")
        return dod_impacts
    
    def _calculate_dod_impact_total(self) -> float:
        """Calculate total DoD impact from all loaded DoD impacts"""
        if not hasattr(self, 'dod_impacts') or not self.dod_impacts:
            print("⚠️  No DoD impacts loaded, using 0% impact")
            return 0.0
        
        total_impact = sum(self.dod_impacts.values())
        print(f"📊 Calculated total DoD impact: {total_impact:.1%} ({total_impact:.3f})")
        return total_impact
    
    def determine_risk_profile(self, story: Dict) -> str:
        """Determine risk profile based on Type of work field with priority system"""
        
        # First, check the Type of work custom field
        type_of_work_value = story.get('fields', {}).get(self.type_of_work_field)
        
        if type_of_work_value:
            # Handle both single values and arrays
            if isinstance(type_of_work_value, list):
                work_types = [item.get('value', '').lower() for item in type_of_work_value]
            elif isinstance(type_of_work_value, dict):
                work_types = [type_of_work_value.get('value', '').lower()]
            elif isinstance(type_of_work_value, str):
                # Handle comma-separated values or single value
                work_types = [wt.strip().lower() for wt in type_of_work_value.split(',')]
            else:
                work_types = []
            
            # Find the highest priority (most risky) type
            highest_priority = 0
            selected_risk = 'experimental'  # default
            
            for work_type in work_types:
                # Check for exact matches or partial matches
                for risk_type, priority in self.risk_priority.items():
                    if risk_type in work_type or work_type in risk_type:
                        if priority > highest_priority:
                            highest_priority = priority
                            selected_risk = risk_type if risk_type != 'dependent' else 'dependant'
            
            if highest_priority > 0:
                print(f"   🏷️  Type of work: {', '.join(work_types)} → {selected_risk}")
                return selected_risk
        
        # Fallback 1: Check labels for risk indicators (legacy support)
        labels = story.get('fields', {}).get('labels', [])
        risk_labels = [label.lower() for label in labels]
        
        if any(label in risk_labels for label in ['proven', 'low-risk', 'fixed']):
            return 'proven'
        elif any(label in risk_labels for label in ['experimental', 'moderate-risk', 'research']):
            return 'experimental'
        elif any(label in risk_labels for label in ['dependant', 'dependent', 'high-risk', 'external']):
            return 'dependant'
        
        # Fallback 2: Default risk assessment based on story points using configurable thresholds
        story_points = self.jira_client.get_story_points(story)
        if story_points:
            risk_config = self.settings_config["risk_assessment"]
            if story_points <= risk_config["proven_threshold_story_points"]:
                return 'proven'
            elif story_points <= risk_config["experimental_threshold_story_points"]:
                return 'experimental'
            else:
                return 'dependant'
        
        print(f"   ⚠️  No Type of work field, using default: experimental")
        return 'experimental'  # Default to experimental
    
    def calculate_prices(self, story_points: float, risk_profile: str) -> Dict[str, float]:
        """Calculate prices based on risk profile and story points using configurable settings"""
        base_price = story_points * self.base_story_point_price
        
        # Apply DoD impact
        base_price_with_dod = base_price * (1 + self.dod_impact_total)
        
        prices = {'base': base_price, 'base_with_dod': base_price_with_dod}
        
        if risk_profile == 'proven':
            prices['fixed'] = base_price_with_dod
            
        elif risk_profile == 'experimental':
            variance = self.experimental_variance
            prices['minimum'] = base_price_with_dod * (1 - variance)
            prices['maximum'] = base_price_with_dod * (1 + variance)
            
        elif risk_profile == 'dependant':
            # Estimate based on hourly rate using configurable hours per story point
            estimated_hours = story_points * self.settings_config["sprint_planning"]["hours_per_story_point"]
            prices['hourly_estimate'] = estimated_hours * self.hourly_rate
        
        return prices
    
    def get_moscow_priority(self, story: Dict) -> str:
        """Determine MoSCoW priority from Jira story"""
        # Check labels first for explicit MoSCoW indicators (higher priority)
        labels = story.get('fields', {}).get('labels', [])
        moscow_labels = [label.lower() for label in labels]
        
        if any(label in moscow_labels for label in ['must', 'must-have']):
            return 'Must Have'
        elif any(label in moscow_labels for label in ['should', 'should-have']):
            return 'Should Have'
        elif any(label in moscow_labels for label in ['could', 'could-have']):
            return 'Could Have'
        elif any(label in moscow_labels for label in ['wont', 'wont-have', 'out-of-scope']):
            return "Won't Have"
        
        # Fall back to priority field mapping if no explicit labels
        priority = story.get('fields', {}).get('priority', {})
        if priority:
            priority_name = priority.get('name', '').lower()
            
            # Map Jira priorities to MoSCoW
            if priority_name in ['highest', 'blocker']:
                return 'Must Have'
            elif priority_name in ['high', 'major']:
                return 'Should Have'
            elif priority_name in ['medium', 'normal']:
                return 'Could Have'
            elif priority_name in ['low', 'minor', 'trivial']:
                return "Won't Have"
        
        return 'Should Have'  # Default

    def get_moscow_priorities_interactive(self) -> List[str]:
        """Interactive method to select which MoSCoW priorities to include"""
        available_priorities = ['Must Have', 'Should Have', 'Could Have', "Won't Have"]
        
        print("\n📋 MoSCoW Priority Filter:")
        print("Select which priorities to include in your spec sheet:")
        print("(You can select multiple priorities)")
        print()
        
        for i, priority in enumerate(available_priorities, 1):
            print(f"{i}. {priority}")
        print("5. All priorities (no filter)")
        print("0. Cancel")
        
        while True:
            try:
                choice = input(f"\nEnter your selection (1-5, or comma-separated for multiple): ").strip()
                
                if choice == "0":
                    print("❌ Operation cancelled.")
                    return None
                
                if choice == "5":
                    print("📊 Including all MoSCoW priorities")
                    return available_priorities
                
                # Handle comma-separated choices
                choices = [int(c.strip()) for c in choice.split(',')]
                selected_priorities = []
                
                for choice_num in choices:
                    if 1 <= choice_num <= 4:
                        priority = available_priorities[choice_num - 1]
                        if priority not in selected_priorities:
                            selected_priorities.append(priority)
                    else:
                        print(f"❌ Invalid choice: {choice_num}. Please select 1-5.")
                        break
                else:
                    # All choices were valid
                    if selected_priorities:
                        print(f"📊 Selected priorities: {', '.join(selected_priorities)}")
                        return selected_priorities
                    else:
                        print("❌ No priorities selected. Please select at least one.")
                
            except ValueError:
                print("❌ Please enter valid numbers separated by commas.")
            except KeyboardInterrupt:
                print("\n⏹️  Selection cancelled. Using all priorities.")
                return available_priorities

    def filter_stories_by_moscow(self, stories: List[Dict], selected_priorities: List[str]) -> tuple[List[Dict], Dict[str, int]]:
        """Filter stories based on selected MoSCoW priorities"""
        if not selected_priorities:
            return stories, {}
        
        filtered_stories = []
        priority_counts = {priority: 0 for priority in ['Must Have', 'Should Have', 'Could Have', "Won't Have"]}
        filtered_counts = {priority: 0 for priority in selected_priorities}
        
        for story in stories:
            moscow_priority = self.get_moscow_priority(story)
            priority_counts[moscow_priority] += 1
            
            if moscow_priority in selected_priorities:
                filtered_stories.append(story)
                filtered_counts[moscow_priority] += 1
        
        # Report filtering results
        total_stories = len(stories)
        filtered_total = len(filtered_stories)
        
        if filtered_total < total_stories:
            print(f"   🔍 Filtered: {filtered_total}/{total_stories} stories match selected priorities")
            for priority in selected_priorities:
                if filtered_counts[priority] > 0:
                    print(f"      - {priority}: {filtered_counts[priority]} stories")
        
        return filtered_stories, priority_counts
    
    def sync_to_scope_sheet(self, sheet_name: str = 'Scope (Quantity)', selected_priorities: List[str] = None):
        """Sync Jira data to the specified scope sheet with epic-story hierarchy plus Everything Else section"""
        print(f"🔄 Regenerating sheet: {sheet_name} from Jira data")
        
        # Use pre-selected epics and everything else items if available
        if hasattr(self, 'selected_epics') and self.selected_epics:
            epics = self.selected_epics
            everything_else_items = getattr(self, 'everything_else_items', [])
            version_info = getattr(self, 'selected_version', 'Selected Version')
            print(f"📋 Using {len(epics)} epic(s) + {len(everything_else_items)} loose items from '{version_info}'")
        else:
            # Fall back to getting epics without version filter
            epics = self.jira_client.get_epics()
            everything_else_items = []
            version_info = "All Epics"
            print(f"📋 Using all epics from project")
            
        if not epics and not everything_else_items:
            print("⚠️  No epics or other items found")
            return False
        
        # Store MoSCoW filter info
        self.selected_moscow_priorities = selected_priorities
        filter_info = f" (Filtered: {', '.join(selected_priorities)})" if selected_priorities and len(selected_priorities) < 4 else ""
        
        # Load the worksheet
        if sheet_name not in self.workbook.sheetnames:
            print(f"❌ Sheet '{sheet_name}' not found in spec sheet")
            return False
        
        ws = self.workbook[sheet_name]
        
        # Clear existing content but preserve the header structure
        self._clear_and_setup_sheet(ws)
        
        # Start adding data after headers (row 4 to account for header structure)
        current_row = 4
        
        print(f"📝 Starting fresh sync at row {current_row}{filter_info}")
        
        total_items = 0
        total_filtered_out = 0
        
        # Process epics and their stories
        print(f"\n📊 Processing {len(epics)} epics...")
        for epic in epics:
            epic_key = epic['key']
            epic_summary = epic['fields']['summary']
            
            print(f"\n🔵 Epic: {epic_key} - {epic_summary}")
            
            # Get stories for this epic
            stories = self.jira_client.get_stories_for_epic(epic_key)
            
            if not stories:
                print(f"   ⚠️  No stories found for epic {epic_key}")
                continue
            
            # Filter stories by MoSCoW priorities if specified
            if selected_priorities and len(selected_priorities) < 4:
                filtered_stories, priority_counts = self.filter_stories_by_moscow(stories, selected_priorities)
                filtered_out_count = len(stories) - len(filtered_stories)
                total_filtered_out += filtered_out_count
            else:
                filtered_stories = stories
            
            if not filtered_stories:
                print(f"   ⚠️  No stories match selected priorities for epic {epic_key}")
                continue
            
            # Add epic header
            self._add_epic_header(ws, current_row, epic_key, epic_summary)
            current_row += 1
            
            # Add stories under this epic
            for story in filtered_stories:
                story_key = story['key']
                story_summary = story['fields']['summary']
                story_points = self.jira_client.get_story_points(story) or 0
                
                # Determine risk profile and pricing
                risk_profile = self.determine_risk_profile(story)
                moscow_priority = self.get_moscow_priority(story)
                prices = self.calculate_prices(story_points, risk_profile)
                
                print(f"   - {story_key}: {story_points} pts, {risk_profile}, {moscow_priority}")
                
                # Add story row
                self._add_story_row(ws, current_row, {
                    'key': story_key,
                    'summary': story_summary,
                    'story_points': story_points,
                    'moscow': moscow_priority,
                    'risk_profile': risk_profile,
                    'prices': prices
                })
                
                current_row += 1
                total_items += 1
            
            # Add spacing after each epic
            current_row += 1
        
        # Process "Everything Else" items not tied to epics
        if everything_else_items:
            print(f"\n📊 Processing {len(everything_else_items)} loose items...")
            
            # Filter everything else items by MoSCoW priorities if specified
            if selected_priorities and len(selected_priorities) < 4:
                filtered_everything_else, priority_counts = self.filter_stories_by_moscow(everything_else_items, selected_priorities)
                filtered_out_count = len(everything_else_items) - len(filtered_everything_else)
                total_filtered_out += filtered_out_count
            else:
                filtered_everything_else = everything_else_items
            
            if filtered_everything_else:
                # Add "Everything Else" section header
                self._add_section_header(ws, current_row, "Everything Else (Not tied to epics)")
                current_row += 1
                
                for item in filtered_everything_else:
                    item_key = item['key']
                    item_summary = item['fields']['summary']
                    story_points = self.jira_client.get_story_points(item) or 0
                    item_type = item['fields'].get('issuetype', {}).get('name', 'Issue')
                    
                    # Determine risk profile and pricing
                    risk_profile = self.determine_risk_profile(item)
                    moscow_priority = self.get_moscow_priority(item)
                    prices = self.calculate_prices(story_points, risk_profile)
                    
                    print(f"   - {item_key} ({item_type}): {story_points} pts, {risk_profile}, {moscow_priority}")
                    
                    # Add item row with type prefix
                    self._add_story_row(ws, current_row, {
                        'key': item_key,
                        'summary': f"[{item_type}] {item_summary}",
                        'story_points': story_points,
                        'moscow': moscow_priority,
                        'risk_profile': risk_profile,
                        'prices': prices
                    })
                    
                    current_row += 1
                    total_items += 1
                
                # Add spacing after everything else section
                current_row += 1
        
        # Add summary section at the end
        if total_items > 0:
            self._add_summary_section(ws, current_row + 1, epics)
        
        # Save the workbook
        self.workbook.save(self.spec_sheet_path)
        
        # Report results
        filter_summary = f" (filtered out {total_filtered_out} items)" if total_filtered_out > 0 else ""
        print(f"\n✅ Sync completed! Generated {total_items} items in {sheet_name}{filter_summary}")
        
        return True
    
    def _clear_and_setup_sheet(self, ws):
        """Clear existing content and set up the sheet with proper headers"""
        # Clear all existing content
        ws.delete_rows(1, ws.max_row)
        
        # Set up the header structure matching the original format
        self._setup_original_headers(ws)
        
        print("🧹 Cleared existing content and set up fresh headers")
    
    def _setup_original_headers(self, ws):
        """Set up the original spec sheet header structure"""
        # Main header row 1 - complex header with scope description
        scope_header = ("Scope\n\nWhat is in scope (features, functionality, integrations) and what is out of scope. "
                       "We define risk profiles transparently. This visibility helps to allocate resources effectively, "
                       "avoid unmanageable scope creep, and ultimately create a shared commitment.\n\n"
                       "* Prices are all-in, meaning they include project management and end-to-end delivery of the product. "
                       "Prices are ex. VAT.")
        
        ws.cell(row=1, column=1).value = scope_header
        ws.cell(row=1, column=2).value = "MoSCoW\n\nMust Have\nShould Have\nCould Have\nWon't Have (out of scope)"
        ws.cell(row=1, column=3).value = "Risk Profile\n\nLow (Proven)\nModerate (Experimental)\nHigh (Dependant)"
        ws.cell(row=1, column=4).value = "Notes\n\n"
        ws.cell(row=1, column=5).value = "Proven\nLow risk profile.\n\nPrice is fixed. Scope is clear and we own this work."
        ws.cell(row=1, column=6).value = "Price\nFixed"
        ws.cell(row=1, column=7).value = "Experimental\nModerate risk profile.\n\nPrice can be 30% lower or higher but never more. We share the risk when the work is experimental."
        ws.cell(row=1, column=8).value = "Price\nMinimum"
        ws.cell(row=1, column=9).value = "Price\nMaximum"
        ws.cell(row=1, column=10).value = "Dependant\nHigh risk profile.\n\nPrice is estimation, but will be pay-by-hour. We depend on to-be-delivered APIs or externally employed team members."
        ws.cell(row=1, column=11).value = "Price\nEstimate, but will be pay-by-hour.\n\nOur rate on 05-02-2025 is €127,16/h - 25% discount."
        
        # Style the headers to match original format
        header_font = Font(bold=True, size=9)
        for col in range(1, 12):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            # Wrap text for better display
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set appropriate column widths from configuration
        column_widths = self.settings_config["ui_formatting"]["default_column_widths"]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Set row height for header from configuration
        ws.row_dimensions[1].height = self.settings_config["ui_formatting"]["header_row_height"]
    
    def _add_summary_section(self, ws, start_row, epics):
        """Add a summary section with totals at the end"""
        current_row = start_row + 1
        
        # Add summary header with version and filter info
        version_info = getattr(self, 'selected_version', 'All Epics') if hasattr(self, 'selected_version') else 'All Epics'
        moscow_filter = getattr(self, 'selected_moscow_priorities', None)
        
        if moscow_filter and len(moscow_filter) < 4:
            filter_text = f" ({', '.join(moscow_filter)} only)"
        else:
            filter_text = ""
            
        ws.cell(row=current_row, column=1).value = f"PROJECT SUMMARY - {version_info}{filter_text}"
        summary_font = Font(bold=True, size=12, color="FFFFFF")
        header_color = self.settings_config["ui_formatting"]["summary_header_color"]
        summary_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        
        for col in range(1, 12):
            cell = ws.cell(row=current_row, column=col)
            cell.font = summary_font
            cell.fill = summary_fill
        
        current_row += 2
        
        # Calculate totals
        total_epics = len(epics)
        total_stories = sum(len(self.jira_client.get_stories_for_epic(epic['key'])) for epic in epics)
        total_story_points = 0
        total_proven_cost = 0
        total_experimental_min = 0
        total_experimental_max = 0
        total_dependant_estimate = 0
        
        for epic in epics:
            stories = self.jira_client.get_stories_for_epic(epic['key'])
            for story in stories:
                story_points = self.jira_client.get_story_points(story) or 0
                total_story_points += story_points
                
                risk_profile = self.determine_risk_profile(story)
                prices = self.calculate_prices(story_points, risk_profile)
                
                if risk_profile == 'proven':
                    total_proven_cost += prices.get('fixed', 0)
                elif risk_profile == 'experimental':
                    total_experimental_min += prices.get('minimum', 0)
                    total_experimental_max += prices.get('maximum', 0)
                elif risk_profile == 'dependant':
                    total_dependant_estimate += prices.get('hourly_estimate', 0)
        
        # Add summary data
        ws.cell(row=current_row, column=1).value = f"Total Epics: {total_epics}"
        current_row += 1
        ws.cell(row=current_row, column=1).value = f"Total Stories: {total_stories}"
        current_row += 1
        ws.cell(row=current_row, column=1).value = f"Total Story Points: {total_story_points}"
        current_row += 2
        
        ws.cell(row=current_row, column=1).value = "COST SUMMARY:"
        current_row += 1
        
        if total_proven_cost > 0:
            ws.cell(row=current_row, column=1).value = "Proven (Fixed):"
            ws.cell(row=current_row, column=6).value = f"€{total_proven_cost:,.2f}"
            current_row += 1
        
        if total_experimental_min > 0:
            ws.cell(row=current_row, column=1).value = "Experimental (Range):"
            ws.cell(row=current_row, column=8).value = f"€{total_experimental_min:,.2f}"
            ws.cell(row=current_row, column=9).value = f"€{total_experimental_max:,.2f}"
            current_row += 1
        
        if total_dependant_estimate > 0:
            ws.cell(row=current_row, column=1).value = "Dependant (Estimate):"
            ws.cell(row=current_row, column=11).value = f"€{total_dependant_estimate:,.2f}"
            current_row += 1
        
        # Add grand total estimate
        grand_total = total_proven_cost + total_experimental_max + total_dependant_estimate
        current_row += 1
        ws.cell(row=current_row, column=1).value = "ESTIMATED TOTAL (worst case):"
        ws.cell(row=current_row, column=11).value = f"€{grand_total:,.2f}"
        
        # Style the totals
        total_font = Font(bold=True, size=11)
        for row in range(start_row + 3, current_row + 1):
            ws.cell(row=row, column=1).font = total_font
    
    def _add_epic_header(self, ws, row, epic_key, epic_summary):
        """Add an epic header row"""
        ws.cell(row=row, column=1).value = f"[EPIC] {epic_summary} ({epic_key})"
        
        # Style the epic header
        epic_font = Font(bold=True, color="000000")
        epic_color = self.settings_config["ui_formatting"]["epic_background_color"]
        epic_fill = PatternFill(start_color=epic_color, end_color=epic_color, fill_type="solid")
        
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.font = epic_font
            cell.fill = epic_fill
    
    def _add_story_row(self, ws, row, story_data):
        """Add a story row with all pricing information"""
        # Column mapping based on the original structure
        ws.cell(row=row, column=1).value = f"  └─ {story_data['summary']} ({story_data['key']})"
        ws.cell(row=row, column=2).value = story_data['moscow']
        ws.cell(row=row, column=3).value = story_data['risk_profile'].title()
        ws.cell(row=row, column=4).value = f"Story Points: {story_data['story_points']}"
        
        # Add pricing based on risk profile
        prices = story_data['prices']
        
        if story_data['risk_profile'] == 'proven':
            ws.cell(row=row, column=5).value = story_data['story_points']  # Story points
            ws.cell(row=row, column=6).value = prices.get('fixed', 0)  # Fixed price
            
        elif story_data['risk_profile'] == 'experimental':
            ws.cell(row=row, column=7).value = story_data['story_points']  # Story points
            ws.cell(row=row, column=8).value = prices.get('minimum', 0)  # Min price
            ws.cell(row=row, column=9).value = prices.get('maximum', 0)  # Max price
            
        elif story_data['risk_profile'] == 'dependant':
            ws.cell(row=row, column=10).value = story_data['story_points']  # Story points
            ws.cell(row=row, column=11).value = prices.get('hourly_estimate', 0)  # Hourly estimate
    
    def _add_section_header(self, ws, row, section_title):
        """Add a section header row for non-epic hierarchies"""
        ws.cell(row=row, column=1).value = f"[SECTION] {section_title}"
        
        # Style the section header
        section_font = Font(bold=True, color="000000")
        section_color = self.settings_config["ui_formatting"]["section_background_color"]
        section_fill = PatternFill(start_color=section_color, end_color=section_color, fill_type="solid")
        
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.font = section_font
            cell.fill = section_fill

    def test_connections(self) -> bool:
        """Test connections to Jira and spec sheet"""
        print("🔄 Testing connections...")
        
        # Test Jira connection
        if not self.jira_client.test_connection():
            return False
        
        # Test spec sheet access
        if not self.load_spec_sheet_structure():
            return False
        
        print("✅ All connections successful")
        return True
    
    def sync_all_sheets(self, selected_priorities: List[str] = None):
        """Sync to all sheets - Jira as single source of truth"""
        # Sync the main scope sheet
        sheet_name = 'Scope (Quantity)'
        
        if sheet_name in self.workbook.sheetnames:
            print(f"\n🔄 Regenerating: {sheet_name}")
            success = self.sync_to_scope_sheet(sheet_name, selected_priorities)
            
            # Add sprint planning analysis if main sync successful
            if success and self.sprint_planning_enabled:
                try:
                    print(f"\n🔄 Adding sprint planning analysis...")
                    self.add_sprint_planning_sheet()
                    print("✅ Sprint planning analysis added")
                except Exception as e:
                    print(f"⚠️  Warning: Could not add sprint planning sheet: {e}")
            

                
        else:
            print(f"❌ Sheet '{sheet_name}' not found in workbook")
    


def main():
    """Main entry point"""
    print("🚀 Enhanced Jira to Spec Sheet Sync Tool")
    print("=" * 60)
    
    sync = EnhancedSpecSheetSync()
    
    # Test connections first
    if not sync.test_connections():
        print("\n❌ Connection test failed. Please check your configuration.")
        sys.exit(1)
    
    # Allow user to select version/release and get epics
    print("\n🔍 Select Version/Release to get Epics:")
    try:
        epics, selected_version = sync.jira_client.get_epics_by_version_interactive()
        
        if not epics:
            print(f"⚠️  No epics found for '{selected_version}'. Please check your selection.")
            return
        
        print(f"\n📊 Found {len(epics)} epic(s) for '{selected_version}'")
        
        # Get all issues for this version to find "Everything Else" items
        print(f"\n🔍 Finding additional items not tied to epics...")
        all_version_issues = sync.jira_client.get_issues_for_version(
            selected_version if selected_version != "All Epics" else None,
            ['Story', 'Bug', 'Task', 'Subtask']  # Include various issue types
        )
        
        # Collect all story keys that are already linked to epics
        linked_story_keys = set()
        for epic in epics:
            epic_stories = sync.jira_client.get_stories_for_epic(epic['key'])
            for story in epic_stories:
                linked_story_keys.add(story['key'])
        
        # Find items not linked to any epic
        everything_else_items = []
        for issue in all_version_issues:
            if issue['key'] not in linked_story_keys:
                everything_else_items.append(issue)
        
        print(f"📋 Found {len(everything_else_items)} additional items not tied to epics")
        
        # Store selected data for the sync
        sync.selected_epics = epics
        sync.selected_version = selected_version
        sync.everything_else_items = everything_else_items
        
    except KeyboardInterrupt:
        print("\n⏹️  Operation cancelled.")
        return
    
    # Allow user to select MoSCoW priorities to include
    print("\n🎯 Filter by MoSCoW Priorities:")
    try:
        selected_priorities = sync.get_moscow_priorities_interactive()
        
        if selected_priorities is None:
            print("⏹️  Operation cancelled.")
            return
            
    except KeyboardInterrupt:
        print("\n⏹️  Selection cancelled. Using all priorities.")
        selected_priorities = ['Must Have', 'Should Have', 'Could Have', "Won't Have"]
    
    # Perform sync
    try:
        sync.sync_all_sheets(selected_priorities)
        priority_filter = f" with {', '.join(selected_priorities)} priorities" if len(selected_priorities) < 4 else ""
        print(f"\n🎉 All done! Synced '{selected_version}'{priority_filter} to your spec sheet.")
    except Exception as e:
        print(f"\n💥 Sync failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main() 