#!/usr/bin/env python3
"""
Enhanced Jira to Spec Sheet Sync Tool
Syncs Jira epics and stories with the existing sophisticated pricing structure
"""

import sys
import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from config import JiraConfig, SpreadsheetConfig
from jira_client import JiraClient
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class EnhancedSpecSheetSync:
    """Enhanced class for synchronizing Jira with the sophisticated spec sheet structure"""
    
    def __init__(self):
        try:
            self.jira_config = JiraConfig()
            self.jira_client = JiraClient(self.jira_config)
            
            # Load existing spec sheet structure
            self.spec_sheet_path = "spec-sheet.xlsx"
            self.base_story_point_price = 130  # €130 per story point
            self.experimental_variance = 0.3  # 30% variance
            self.hourly_rate = 127.16 * 0.75  # €127.16/h with 25% discount
            self.dod_impact_total = 0.63  # 63% total DoD impact
            
        except Exception as e:
            print(f"❌ Configuration error: {e}")
            sys.exit(1)
    
    def load_spec_sheet_structure(self):
        """Load the existing spec sheet or create a new one"""
        try:
            if os.path.exists(self.spec_sheet_path):
                self.workbook = openpyxl.load_workbook(self.spec_sheet_path)
                
                # Load DoD impacts
                self.dod_impacts = self._load_dod_impacts()
                
                # Load settings
                self.settings = self._load_settings()
                
                print("✅ Loaded existing spec sheet structure")
            else:
                print(f"📝 Spec sheet not found at {self.spec_sheet_path}, creating new one...")
                self._create_new_spec_sheet()
                print("✅ Created new spec sheet with default structure")
            
            return True
            
        except Exception as e:
            print(f"❌ Error with spec sheet: {e}")
            return False
    
    def _load_dod_impacts(self) -> Dict[str, float]:
        """Load Definition of Done impacts from the spec sheet"""
        dod_impacts = {}
        try:
            df = pd.read_excel(self.spec_sheet_path, sheet_name='Definition of Done (Quality)')
            
            for idx, row in df.iterrows():
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[3]):
                    description = str(row.iloc[0])
                    impact = float(row.iloc[3])
                    if impact > 0:
                        dod_impacts[description] = impact
                        
        except Exception as e:
            print(f"Warning: Could not load DoD impacts: {e}")
            
        return dod_impacts
    
    def _load_settings(self) -> Dict[str, any]:
        """Load settings from the spec sheet"""
        settings = {
            'base_story_point_price': 130,
            'experimental_variance': 0.3,
            'hourly_rate': 95.37  # 127.16 * 0.75
        }
        
        try:
            df = pd.read_excel(self.spec_sheet_path, sheet_name='Settings')
            
            for idx, row in df.iterrows():
                if pd.notna(row.iloc[0]):
                    setting_name = str(row.iloc[0]).lower()
                    if 'base price of 1 story point' in setting_name and pd.notna(row.iloc[1]):
                        settings['base_story_point_price'] = float(row.iloc[1])
                    elif 'experimental' in setting_name and 'range' in setting_name and pd.notna(row.iloc[1]):
                        settings['experimental_variance'] = float(row.iloc[1])
                        
        except Exception as e:
            print(f"Warning: Could not load settings: {e}")
            
        return settings
    
    def _create_new_spec_sheet(self):
        """Create a new spec sheet with all required sheets"""
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Create Scope (Quantity) sheet
        scope_sheet = self.workbook.create_sheet('Scope (Quantity)')
        
        # Create Definition of Done sheet with default values
        dod_sheet = self.workbook.create_sheet('Definition of Done (Quality)')
        self._setup_default_dod_sheet(dod_sheet)
        
        # Create Settings sheet with default values
        settings_sheet = self.workbook.create_sheet('Settings')
        self._setup_default_settings_sheet(settings_sheet)
        
        # Set default values
        self.dod_impacts = self._get_default_dod_impacts()
        self.settings = self._load_settings()
        
        # Save the new workbook
        self.workbook.save(self.spec_sheet_path)
    
    def _setup_default_dod_sheet(self, ws):
        """Set up default Definition of Done sheet"""
        headers = [
            "Definition of Done",
            "MoSCoW",
            "Price Impact",
            "Price Impact %"
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i).value = header
            ws.cell(row=1, column=i).font = Font(bold=True)
        
        # Add some default DoD items
        dod_items = [
            ("Code quality & documentation", "Must Have", "", 0.04),
            ("Code is reviewed and approved via pull requests", "Must Have", "", 0.025),
            ("Code is well-documented", "Should Have", "", 0.04),
            ("Performance optimization", "Should Have", "", 0.065),
            ("Testing & Cross-browser compatibility", "Must Have", "", 0.065),
            ("Security vulnerabilities are identified", "Must Have", "", 0.04),
        ]
        
        for i, (desc, moscow, impact_desc, impact_pct) in enumerate(dod_items, 2):
            ws.cell(row=i, column=1).value = desc
            ws.cell(row=i, column=2).value = moscow
            ws.cell(row=i, column=3).value = impact_desc
            ws.cell(row=i, column=4).value = impact_pct
    
    def _setup_default_settings_sheet(self, ws):
        """Set up default Settings sheet"""
        settings_data = [
            ("Range for Experimental pricing (up & down)", 0.3),
            ("Base price of 8 story points", 1040),
            ("Base price of 1 story point", 130),
            ("Hourly rate", 95.37),
        ]
        
        ws.cell(row=1, column=1).value = "Settings"
        ws.cell(row=1, column=1).font = Font(bold=True)
        
        for i, (setting, value) in enumerate(settings_data, 2):
            ws.cell(row=i, column=1).value = setting
            ws.cell(row=i, column=2).value = value
    
    def _get_default_dod_impacts(self):
        """Get default DoD impacts when no sheet exists"""
        return {
            "code_quality": 0.04,
            "code_review": 0.025,
            "documentation": 0.04,
            "performance": 0.065,
            "testing": 0.065,
            "security": 0.04
        }
    
    def determine_risk_profile(self, story: Dict) -> str:
        """Determine risk profile based on Jira story labels or custom fields"""
        # Check labels for risk indicators
        labels = story.get('fields', {}).get('labels', [])
        
        # Check for risk-related labels
        risk_labels = [label.lower() for label in labels]
        
        if any(label in risk_labels for label in ['proven', 'low-risk', 'fixed']):
            return 'proven'
        elif any(label in risk_labels for label in ['experimental', 'moderate-risk', 'research']):
            return 'experimental'
        elif any(label in risk_labels for label in ['dependant', 'dependent', 'high-risk', 'external']):
            return 'dependant'
        
        # Default risk assessment based on story points
        story_points = self.jira_client.get_story_points(story)
        if story_points:
            if story_points <= 3:
                return 'proven'
            elif story_points <= 8:
                return 'experimental'
            else:
                return 'dependant'
        
        return 'experimental'  # Default to experimental
    
    def calculate_prices(self, story_points: float, risk_profile: str) -> Dict[str, float]:
        """Calculate prices based on risk profile and story points"""
        base_price = story_points * self.settings['base_story_point_price']
        
        # Apply DoD impact
        base_price_with_dod = base_price * (1 + self.dod_impact_total)
        
        prices = {'base': base_price, 'base_with_dod': base_price_with_dod}
        
        if risk_profile == 'proven':
            prices['fixed'] = base_price_with_dod
            
        elif risk_profile == 'experimental':
            variance = self.settings['experimental_variance']
            prices['minimum'] = base_price_with_dod * (1 - variance)
            prices['maximum'] = base_price_with_dod * (1 + variance)
            
        elif risk_profile == 'dependant':
            # Estimate based on hourly rate
            estimated_hours = story_points * 8  # Rough estimate: 8 hours per story point
            prices['hourly_estimate'] = estimated_hours * self.settings['hourly_rate']
        
        return prices
    
    def get_moscow_priority(self, story: Dict) -> str:
        """Determine MoSCoW priority from Jira story"""
        priority = story.get('fields', {}).get('priority', {})
        if priority:
            priority_name = priority.get('name', '').lower()
            
            # Map Jira priorities to MoSCoW
            if priority_name in ['highest', 'blocker']:
                return 'Must Have'
            elif priority_name in ['high', 'major']:
                return 'Should Have'
            elif priority_name in ['medium', 'normal']:
                return 'Could Have'
            elif priority_name in ['low', 'minor', 'trivial']:
                return "Won't Have"
        
        # Check labels for MoSCoW indicators
        labels = story.get('fields', {}).get('labels', [])
        moscow_labels = [label.lower() for label in labels]
        
        if any(label in moscow_labels for label in ['must', 'must-have']):
            return 'Must Have'
        elif any(label in moscow_labels for label in ['should', 'should-have']):
            return 'Should Have'
        elif any(label in moscow_labels for label in ['could', 'could-have']):
            return 'Could Have'
        elif any(label in moscow_labels for label in ['wont', 'wont-have', 'out-of-scope']):
            return "Won't Have"
        
        return 'Should Have'  # Default
    
    def sync_to_scope_sheet(self, sheet_name: str = 'Scope (Quantity)'):
        """Sync Jira data to the specified scope sheet - completely regenerate from Jira"""
        print(f"🔄 Regenerating sheet: {sheet_name} from Jira data")
        
        # Get Jira data
        epics = self.jira_client.get_epics()
        if not epics:
            print("⚠️  No epics found in Jira project")
            return False
        
        # Load the worksheet
        if sheet_name not in self.workbook.sheetnames:
            print(f"❌ Sheet '{sheet_name}' not found in spec sheet")
            return False
        
        ws = self.workbook[sheet_name]
        
        # Clear existing content but preserve the header structure
        self._clear_and_setup_sheet(ws)
        
        # Start adding data after headers (row 4 to account for header structure)
        current_row = 4
        
        print(f"📝 Starting fresh sync at row {current_row}")
        
        total_items = 0
        
        for epic in epics:
            epic_key = epic['key']
            epic_summary = epic['fields']['summary']
            
            print(f"\n📊 Processing epic: {epic_key} - {epic_summary}")
            
            # Add epic header
            self._add_epic_header(ws, current_row, epic_key, epic_summary)
            current_row += 1
            
            # Get stories for this epic
            stories = self.jira_client.get_stories_for_epic(epic_key)
            print(f"   Found {len(stories)} stories")
            
            for story in stories:
                story_key = story['key']
                story_summary = story['fields']['summary']
                story_points = self.jira_client.get_story_points(story) or 0
                
                # Determine risk profile and pricing
                risk_profile = self.determine_risk_profile(story)
                moscow_priority = self.get_moscow_priority(story)
                prices = self.calculate_prices(story_points, risk_profile)
                
                print(f"   - {story_key}: {story_points} pts, {risk_profile}, {moscow_priority}")
                
                # Add story row
                self._add_story_row(ws, current_row, {
                    'key': story_key,
                    'summary': story_summary,
                    'story_points': story_points,
                    'moscow': moscow_priority,
                    'risk_profile': risk_profile,
                    'prices': prices
                })
                
                current_row += 1
                total_items += 1
            
            # Add spacing between epics
            current_row += 1
        
        # Add summary row at the end
        if total_items > 0:
            self._add_summary_section(ws, current_row + 1, epics)
        
        # Save the workbook
        self.workbook.save(self.spec_sheet_path)
        print(f"\n✅ Sync completed! Generated {total_items} items in {sheet_name}")
        
        return True
    
    def _clear_and_setup_sheet(self, ws):
        """Clear existing content and set up the sheet with proper headers"""
        # Clear all existing content
        ws.delete_rows(1, ws.max_row)
        
        # Set up the header structure matching the original format
        self._setup_original_headers(ws)
        
        print("🧹 Cleared existing content and set up fresh headers")
    
    def _setup_original_headers(self, ws):
        """Set up the original spec sheet header structure"""
        # Main header row 1 - complex header with scope description
        scope_header = ("Scope\n\nWhat is in scope (features, functionality, integrations) and what is out of scope. "
                       "We define risk profiles transparently. This visibility helps to allocate resources effectively, "
                       "avoid unmanageable scope creep, and ultimately create a shared commitment.\n\n"
                       "* Prices are all-in, meaning they include project management and end-to-end delivery of the product. "
                       "Prices are ex. VAT.")
        
        ws.cell(row=1, column=1).value = scope_header
        ws.cell(row=1, column=2).value = "MoSCoW\n\nMust Have\nShould Have\nCould Have\nWon't Have (out of scope)"
        ws.cell(row=1, column=3).value = "Risk Profile\n\nLow (Proven)\nModerate (Experimental)\nHigh (Dependant)"
        ws.cell(row=1, column=4).value = "Notes\n\n"
        ws.cell(row=1, column=5).value = "Proven\nLow risk profile.\n\nPrice is fixed. Scope is clear and we own this work."
        ws.cell(row=1, column=6).value = "Price\nFixed"
        ws.cell(row=1, column=7).value = "Experimental\nModerate risk profile.\n\nPrice can be 30% lower or higher but never more. We share the risk when the work is experimental."
        ws.cell(row=1, column=8).value = "Price\nMinimum"
        ws.cell(row=1, column=9).value = "Price\nMaximum"
        ws.cell(row=1, column=10).value = "Dependant\nHigh risk profile.\n\nPrice is estimation, but will be pay-by-hour. We depend on to-be-delivered APIs or externally employed team members."
        ws.cell(row=1, column=11).value = "Price\nEstimate, but will be pay-by-hour.\n\nOur rate on 05-02-2025 is €127,16/h - 25% discount."
        
        # Style the headers to match original format
        header_font = Font(bold=True, size=9)
        for col in range(1, 12):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            # Wrap text for better display
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set appropriate column widths
        column_widths = [50, 20, 20, 30, 15, 12, 15, 12, 12, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Set row height for header
        ws.row_dimensions[1].height = 80
    
    def _add_summary_section(self, ws, start_row, epics):
        """Add a summary section with totals at the end"""
        current_row = start_row + 1
        
        # Add summary header
        ws.cell(row=current_row, column=1).value = "PROJECT SUMMARY"
        summary_font = Font(bold=True, size=12, color="FFFFFF")
        summary_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in range(1, 12):
            cell = ws.cell(row=current_row, column=col)
            cell.font = summary_font
            cell.fill = summary_fill
        
        current_row += 2
        
        # Calculate totals
        total_epics = len(epics)
        total_stories = sum(len(self.jira_client.get_stories_for_epic(epic['key'])) for epic in epics)
        total_story_points = 0
        total_proven_cost = 0
        total_experimental_min = 0
        total_experimental_max = 0
        total_dependant_estimate = 0
        
        for epic in epics:
            stories = self.jira_client.get_stories_for_epic(epic['key'])
            for story in stories:
                story_points = self.jira_client.get_story_points(story) or 0
                total_story_points += story_points
                
                risk_profile = self.determine_risk_profile(story)
                prices = self.calculate_prices(story_points, risk_profile)
                
                if risk_profile == 'proven':
                    total_proven_cost += prices.get('fixed', 0)
                elif risk_profile == 'experimental':
                    total_experimental_min += prices.get('minimum', 0)
                    total_experimental_max += prices.get('maximum', 0)
                elif risk_profile == 'dependant':
                    total_dependant_estimate += prices.get('hourly_estimate', 0)
        
        # Add summary data
        ws.cell(row=current_row, column=1).value = f"Total Epics: {total_epics}"
        current_row += 1
        ws.cell(row=current_row, column=1).value = f"Total Stories: {total_stories}"
        current_row += 1
        ws.cell(row=current_row, column=1).value = f"Total Story Points: {total_story_points}"
        current_row += 2
        
        ws.cell(row=current_row, column=1).value = "COST SUMMARY:"
        current_row += 1
        
        if total_proven_cost > 0:
            ws.cell(row=current_row, column=1).value = "Proven (Fixed):"
            ws.cell(row=current_row, column=6).value = f"€{total_proven_cost:,.2f}"
            current_row += 1
        
        if total_experimental_min > 0:
            ws.cell(row=current_row, column=1).value = "Experimental (Range):"
            ws.cell(row=current_row, column=8).value = f"€{total_experimental_min:,.2f}"
            ws.cell(row=current_row, column=9).value = f"€{total_experimental_max:,.2f}"
            current_row += 1
        
        if total_dependant_estimate > 0:
            ws.cell(row=current_row, column=1).value = "Dependant (Estimate):"
            ws.cell(row=current_row, column=11).value = f"€{total_dependant_estimate:,.2f}"
            current_row += 1
        
        # Add grand total estimate
        grand_total = total_proven_cost + total_experimental_max + total_dependant_estimate
        current_row += 1
        ws.cell(row=current_row, column=1).value = "ESTIMATED TOTAL (worst case):"
        ws.cell(row=current_row, column=11).value = f"€{grand_total:,.2f}"
        
        # Style the totals
        total_font = Font(bold=True, size=11)
        for row in range(start_row + 3, current_row + 1):
            ws.cell(row=row, column=1).font = total_font
    
    def _add_epic_header(self, ws, row, epic_key, epic_summary):
        """Add an epic header row"""
        ws.cell(row=row, column=1).value = f"[EPIC] {epic_summary} ({epic_key})"
        
        # Style the epic header
        epic_font = Font(bold=True, color="000000")
        epic_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.font = epic_font
            cell.fill = epic_fill
    
    def _add_story_row(self, ws, row, story_data):
        """Add a story row with all pricing information"""
        # Column mapping based on the original structure
        ws.cell(row=row, column=1).value = f"  └─ {story_data['summary']} ({story_data['key']})"
        ws.cell(row=row, column=2).value = story_data['moscow']
        ws.cell(row=row, column=3).value = story_data['risk_profile'].title()
        ws.cell(row=row, column=4).value = f"Story Points: {story_data['story_points']}"
        
        # Add pricing based on risk profile
        prices = story_data['prices']
        
        if story_data['risk_profile'] == 'proven':
            ws.cell(row=row, column=5).value = story_data['story_points']  # Story points
            ws.cell(row=row, column=6).value = prices.get('fixed', 0)  # Fixed price
            
        elif story_data['risk_profile'] == 'experimental':
            ws.cell(row=row, column=7).value = story_data['story_points']  # Story points
            ws.cell(row=row, column=8).value = prices.get('minimum', 0)  # Min price
            ws.cell(row=row, column=9).value = prices.get('maximum', 0)  # Max price
            
        elif story_data['risk_profile'] == 'dependant':
            ws.cell(row=row, column=10).value = story_data['story_points']  # Story points
            ws.cell(row=row, column=11).value = prices.get('hourly_estimate', 0)  # Hourly estimate
    
    def test_connections(self) -> bool:
        """Test connections to Jira and spec sheet"""
        print("🔄 Testing connections...")
        
        # Test Jira connection
        if not self.jira_client.test_connection():
            return False
        
        # Test spec sheet access
        if not self.load_spec_sheet_structure():
            return False
        
        print("✅ All connections successful")
        return True
    
    def sync_all_sheets(self):
        """Sync to the main scope sheet only - Jira as single source of truth"""
        sheet_name = 'Scope (Quantity)'
        
        if sheet_name in self.workbook.sheetnames:
            print(f"\n🔄 Regenerating: {sheet_name}")
            self.sync_to_scope_sheet(sheet_name)
        else:
            print(f"❌ Sheet '{sheet_name}' not found in workbook")

def main():
    """Main entry point"""
    print("🚀 Enhanced Jira to Spec Sheet Sync Tool")
    print("=" * 60)
    
    sync = EnhancedSpecSheetSync()
    
    # Test connections first
    if not sync.test_connections():
        print("\n❌ Connection test failed. Please check your configuration.")
        sys.exit(1)
    
    # Perform sync
    try:
        sync.sync_all_sheets()
        print("\n🎉 All done! Check your spec sheet for updated Jira data.")
    except Exception as e:
        print(f"\n💥 Sync failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main() 